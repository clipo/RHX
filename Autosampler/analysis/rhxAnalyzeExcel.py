__author__ = 'clipo'

import sys
#sys.path.insert(0, "/Library/Frameworks/Python.framework/Versions/7.2/")
import easygui
import pylab
import numpy
import logging
import time
from ctypes import *
import math
import os
import io
from pylab import *
from time import sleep
import sched, time
import signal
import random
import csv
import re
from datetime import date
import numpy as np
from scipy import stats
from numpy import *
from datetime import datetime
from datetime import timedelta
import os.path
import os
from array import *
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.mlab as mlab
import matplotlib.cbook as cbook
import pylab
import threading
import signal
import time
from scipy import signal
from openpyxl import Workbook
from openpyxl import load_workbook

def smoothListGaussian(list,strippedXs=False,degree=5):
   window=degree*2-1
   weight=numpy.array([1.0]*window)
   weightGauss=[]
   for i in range(window):
       i=i-degree+1
       frac=i/float(window)
       gauss=1/(numpy.exp((4*(frac))**2))
       weightGauss.append(gauss)
   weight=numpy.array(weightGauss)*weight
   smoothed=[0.0]*(len(list)-window)
   for i in range(len(smoothed)):
       smoothed[i]=sum(numpy.array(list[i:i+window])*weight)/sum(weight)
   return smoothed

def evenlySpacedPlot(evenPlot,evenXarray,evenYarray,liquidwater):
   evenPlot.scatter(evenXarray,evenYarray, marker='o', c='r')
   evenPlot.set_xlabel('Time (m^1/4)', fontsize=8)
   evenPlot.set_ylabel('Mass change (% of initial mass)', fontsize=8)
   ymin=min(evenYarray)
   ymax=max(evenYarray)
   evenPlot.set_ylim(ymin,ymax)
   xcoords = np.array([evenXarray[0],evenXarray[len(evenXarray)-1]])
   ycoords =  np.array([liquidwater,liquidwater])
   evenPlot.plot(xcoords,ycoords, color='b', linewidth=1, alpha=0.4)
   evenPlot.grid(True)
   plt.ion()     # turns on interactive mode
   plt.show()    # now this should be non-blocking


def temp_humidity_plot(tempPlot,humidPlot,timeArray,tempArray,humidArray):
   #tempPlot.tick(labelsize=8)
   #humidPlot.tick(labelsize=8)
   tempPlot.scatter(timeArray,tempArray, marker='^', c='r')
   humidPlot.scatter(timeArray,humidArray, marker='o', c='g')
   tempPlot.set_xlabel('Time (m^1/4)', fontsize=8)
   humidPlot.set_xlabel('Time (m^1/4)', fontsize=8)
   tempPlot.set_ylabel('Non-Filtered Temperature (C)', fontsize=8)
   humidPlot.set_ylabel('Non-Filtered Humidity (%RH)', fontsize=8)
   tempPlot.grid(True)
   humidPlot.grid(True)
   plt.ion()     # turns on interactive mode
   plt.show()    # now this should be non-blocking

def correctForRH( xArray, yArray, humidArray, humiditySetting, rhCoefficent ):
   correctedArray=zeros(alen(yArray))
   num=0
   for oldWeight in yArray:
      humidity=humidArray[ num ]
      newWeight=oldWeight+rhCoefficent*(humidity-humiditySetting)/humiditySetting
      correctedArray[num]=newWeight
      num+=1
   return correctedArray


def filtered_temp_humidity_plot(filteredTempPlot,filteredHumidPlot,timeArray,filteredTempArray,filteredHumidArray):
   #filteredTempPlot.tick(labelsize=8)
   filteredTempPlot.scatter(timeArray,filteredTempArray, marker='^', c='r')
   filteredHumidPlot.scatter(timeArray,filteredHumidArray, marker='o', c='g')
   filteredTempPlot.set_xlabel('Time (m^1/4)', fontsize=8)
   filteredHumidPlot.set_xlabel('Time (m^1/4)', fontsize=8)
   filteredTempPlot.set_ylabel('Filtered Temperature (C)', fontsize=8)
   filteredHumidPlot.set_ylabel('Filtered Humidity (%RH)', fontsize=8)
   plt.ion()     # turns on interactive mode
   plt.show()    # now this should be non-blocking


def overall_plot(ax1,xArray,yArray, sampleNumber,liquidwater):
   #ax1.tick(labelsize=8)
   minx=min(xArray)
   miny=min(yArray)
   ymin=min(yArray)
   ymax=max(yArray)
   xmin=min(xArray)
   xmax=min(xArray)
   ax1.scatter(xArray,yArray, marker='x', c='r')
   ax1.set_ylim(ymin,ymax)
   ax1.set_xlabel('Time (m^1/4)', fontsize=8)
   ax1.set_ylabel('Weight (g)', fontsize=8)
   #ax1.tick(labelsize=8)

   #print "minx:", minx
   #label="Sample Number: "+str(sampleNumber)
   #plt.text(minx+.2,miny,label,fontsize=8)
   xcoords = np.array([xArray[0],xArray[len(xArray)-1]])
   ycoords = np.array([liquidwater,liquidwater])
   ax1.plot(xcoords,ycoords, color='b', linewidth=1, alpha=0.4)
   ax1.grid(True)
   plt.ion()     # turns on interactive mode
   plt.show()    # now this should be non-blocking

def linear_plot(ax2,xarray,yarray,minx,maxx,miny,maxy,slope,intercept,p_value,r_value,ffilename):
   #ax2.tick(labelsize=8)
   ax2.plot(xarray, yarray, 'x')
   xls = (minx,maxx)
   yls = (miny,maxy)
   x1=0.0
   x2=0.0
   ax2.axis([xls[0], xls[1], yls[0], yls[1]])
   ax2.set_xlabel('Time (m^1/4)', fontsize=8)
   ax2.set_ylabel('% Weight Change', fontsize=8)
   ax2.ticklabel_format(style='plain', axis='both, fontsize=8')
   #title = ffilename + " Sample Number: " + str(sampleNumber)
   #ax2.set_title(title, fontsize=8)
   ax2.grid(True)
   if slope == 0.0 :# the line is parallel to the x axis
      x1 = xls[0]
      x2 = xls[1]
   elif slope > 0.0 :
      # if m > 0, smallest x is the larger of the x values of the points where the line intersects
      # the x = xmin(xls[0]) and y = ymin(yls[0]) lines
      # and largest x is the smaller of the x values of the points where the line intersects
      # the x = xmax(xls[1]) and y = ymax(yls[1]) lines
      x1 = max(xls[0],(yls[0] - intercept)/slope)
      x2 = min(xls[1],(yls[1] - intercept)/slope)
   else :
      # if m < 0, smallest x is the larger of the x values of the points where the line intersects
      # the x = xmin(xls[0]) and y = ymax(yls[1]) lines
      # and largest x is the smaller of the x values of the points where the line intersects
      # the x = xmax(xls[1]) and y = ymin(yls[0]) lines
      x1 = max(xls[0],(yls[1] - intercept)/slope)
      x2 = min(xls[1],(yls[0] - intercept)/slope)
   xcoords = np.array([x1,x2])
   #because of numpy entire set of xcoords can be processed at once to generate the ycoords
   ycoords = slope*xcoords + intercept

   ax2.plot(xcoords,ycoords, color='r', linewidth=4, label='y = ' + str(slope) + '*x ' +
                                                          '+-'[intercept < 0.0] + ' ' + str(abs(intercept)))

   ax2.grid(True)
   plt.ion()     # turns on interactive mode
   plt.show()    # now this should be non-blocking

def ask_for_RHCoefficient():
   msg         = "Enter Humidity Information"
   title       = "Humidity"
   fieldNames  = ["RH Setting (%RH)","Coefficient for Weight Correlation to RH"]
   fieldValues = []  # we start with blanks for the values
   fieldValues = easygui.multenterbox(msg,title, fieldNames)
   # make sure that none of the fields was left blank
   while 1:  # do forever, until we find acceptable values and break out
      if fieldValues is None:
         break
      errmsg = ""
      if fieldValues[0] <1 or  fieldValues[1] < 1:
         errmsg += ("You must enter a postive number.")
         # look for errors in the returned values
      for i in range(len(fieldNames)):
         if fieldValues[i].strip() == "":
            errmsg +=('"%s" is a required field.\n\n' % fieldNames[i])
      if errmsg == "":
         break # no problems found
      else:
         # show the box again, with the errmsg as the message
         errmsg += " Invalid values. Please try again. "
         fieldValues = easygui.multenterbox(errmsg, title, fieldNames, fieldValues)

   return float(fieldValues[0]),float(fieldValues[1])

def ask_for_time_delay():
      msg         = "Enter time (minutes) since 550C firing ended"

      fieldValue = 0  # we start with blanks for the values
      fieldValue = easygui.enterbox(msg,"0")
      return fieldValue

def ask_for_worksheet(choices):
   msg         = "Choose the worksheet that contains the run data."
   title       = "Worksheet Choice"
   choice = easygui.choicebox(msg, title, choices)
   return choice

def ask_for_weight_values():
   msg         = "Enter Weight Data"
   title       = "Weights"
   fieldNames  = ["Initial Weight (g)","Std Dev","105 Degrees C Weight (g)","Std Dev","550 Degrees C Weight","Std Dev"]
   fieldValues = []  # we start with blanks for the values
   fieldValues = easygui.multenterbox(msg,title, fieldNames)
   # make sure that none of the fields was left blank
   while 1:  # do forever, until we find acceptable values and break out
      if fieldValues is None:
         break
      errmsg = ""
      if float(fieldValues[0]) < float(fieldValues[2]):
         errmsg += ("The initial weight must be higher than the 105 degree weight ")
         # look for errors in the returned values
      if float(fieldValues[2]) < float(fieldValues[4]):
         errmsg += ("The 105 degree weight must be higher than the 550 degree weight ")
         # look for errors in the returned values
      for i in range(len(fieldNames)):
         if fieldValues[i].strip() == "":
            errmsg +=('"%s" is a required field.\n\n' % fieldNames[i])
      if errmsg == "":
         break # no problems found
      else:
         # show the box again, with the errmsg as the message
         errmsg += " Invalid values. Please try again. "
         fieldValues = easygui.multenterbox(errmsg, title, fieldNames, fieldValues)

   return float(fieldValues[0]),float(fieldValues[1]),float(fieldValues[2]),float(fieldValues[3]),float(fieldValues[4]),float(fieldValues[5])

def ask_for_minmax_values(minx,maxx):
   msg         = "Information for linear calculations"
   title       = "Linear Calculations"
   fieldNames  = ["Minimum time (m^1/4)","Maximum time (m^1/4)"]
   fieldValues = []  # we start with blanks for the values
   fieldValues = easygui.multenterbox(msg,title, fieldNames)
   # make sure that none of the fields was left blank
   while 1:  # do forever, until we find acceptable values and break out
      if fieldValues is None:
         break
      errmsg = ""
      if float(fieldValues[0]) > float(fieldValues[1]):
         errmsg += ("The minimum value must be lower than the maximum value. ")
      # look for errors in the returned values
      for i in range(len(fieldNames)):
         if fieldValues[i].strip() == "":
            errmsg += ('"%s" is a required field.\n\n' % fieldNames[i])
      if errmsg == "":
         break # no problems found
      else:
         # show the box again, with the errmsg as the message
         errmsg += " Invalid values. Please try again. "
         fieldValues = easygui.multenterbox(errmsg, title, fieldNames, fieldValues)

   return float(fieldValues[0]),float(fieldValues[1])

def graph_slope(sampleNumber,x,y,window,originalWeightAverage,originalWeightStdDev,prefireWeightAverage,prefireWeightStdDev,postfireWeightAverage,postfireWeightStdDev):
   #window = [-1, 0, 0,0,0,0,0,0, 1]
   #slope = convolve(y, window, mode='same') / convolve(x, window, mode='same')
   # a[start,end,step]

   xs = zeros(alen(x))
   ys = zeros(alen(x))
   slopeArray=zeros(alen(x))
   interceptArray=zeros(alen(x))
   r_valueArray=zeros(alen(x))
   p_valueArray=zeros(alen(x))
   std_errArray=zeros(alen(x))
   date_Array=zeros(alen(x))
   slope=0.0
   intercept=0.0
   r_value=0.0
   p_value=0.0
   std_err=0.0
   date=0.0
   num=1
   dateYears=0.0
   dateCalendar=0.0
   ADBC=""
   dateErrorYear=0.0

   while num < alen(x):
      #print "Num: ", num
      if num<window:
         slopeArray[num]=0.0
         interceptArray[num]=0.0
         r_valueArray[num]=0.0
         p_valueArray[num]=0.0
         std_errArray[num]=0.0
         #print "slope: 0"
      else:
         xs=x[num-window:num:1]
         ys=y[num-window:num:1]
         #print xs
         slope, intercept, r_value, p_value, std_err = stats.linregress(xs,ys)
         #print "slope:", slope, " -- ", num

         slopeArray[num]=slope
         interceptArray[num]=intercept
         r_valueArray[num]=r_value
         p_valueArray[num]=p_value
         std_errArray[num]=std_err
         #print "Slope Array: ", alen(slopeArray)
         #print slopeArray

         dateYears,dateCalendar,ADBC,dateErrorYear=age_calculate(slope,std_err, postfireWeightAverage,
            postfireWeightStdDev,prefireWeightAverage,prefireWeightStdDev)
         date_Array[num]=dateYears
         #print dateYears
      num+=1


   #print slopeArray
      #print "LENGTH: ", alen(slopeArray)
   return (slopeArray,interceptArray,r_valueArray,p_valueArray,std_errArray,date_Array)

def age_calculateFromPercent(pslope,pstd_err,percentWeightChange):

   ## now divide weight by the slope
   dateQuarterPower = percentWeightChange/pslope
   ## now exp to the 4 power
   dateMinutes = pow(abs(dateQuarterPower),4)
   ## 60 minutes per hour
   dateHours = dateMinutes/60
   ## 24 hours to day
   dateDays = dateHours/24
   ## 365 hours/year
   dateYears=dateDays/365
   ## AD/BC - subtract from this year
   ## initialize the maxDate variable with a date (today)
   ADBC=""

   mdate=date.today()

   maxDateYear=mdate.year
   ## now measure the difference-- the year at the end of the measurement minus number of years in calculations
   dateCalendar=maxDateYear - dateYears
   if dateCalendar <0:
      ADBC ="BC"
   else:
      ADBC="AD"

   # now calculate error terms
   # this consists of the prefire weight error, the postfire weight error, the slope error,
   # where prefire-postfire/slope
   # so date * sqrt( ( prefireError/prefire)^2 + (postfireError/postfire)^2 + *slopeError/slope)^2)
   total_error = dateYears * sqrt( pow(pstd_err/pslope, 2))

   return (dateYears,dateCalendar,ADBC,total_error)


def age_calculate(slope,slopeError,postfireWeight,postfireWeightStdDev,prefireWeight,prefireWeightStdDev):

   #print "Prefire weight: ", weightAverage, "+/- ", weightStdDev

   ## difference is the pre fire weight  minus intercept (post fire weight)
   diffWeight = prefireWeight - postfireWeight
   ## now divide weight by the slope
   dateQuarterPower = diffWeight/slope
   ## now exp to the 4 power
   dateMinutes = pow(abs(dateQuarterPower),4)
   ## 60 minutes per hour
   dateHours = dateMinutes/60
   ## 24 hours to day
   dateDays = dateHours/24
   ## 365 hours/year
   dateYears=dateDays/365
   ## AD/BC - subtract from this year
   ## initialize the maxDate variable with a date (today)
   ADBC=""

   mdate=date.today()

   maxDateYear=mdate.year
   ## now measure the difference-- the year at the end of the measurement minus number of years in calculations
   dateCalendar=maxDateYear - dateYears
   if dateCalendar <0:
      ADBC ="BC"
   else:
      ADBC="AD"
   if postfireWeightStdDev is None:
      postfireWeightStdDev=0.0
   dateErrorMinutes=pow(postfireWeightStdDev,4)
   dateErrorHours=dateErrorMinutes/60
   dateErrorDays=dateErrorHours/24
   dateErrorYears=dateErrorDays/365

   # now calculate error terms
   # this consists of the prefire weight error, the postfire weight error, the slope error,
   # where prefire-postfire/slope
   # so date * sqrt( ( prefireError/prefire)^2 + (postfireError/postfire)^2 + *slopeError/slope)^2)
   total_error = dateYears * sqrt( pow(prefireWeightStdDev/prefireWeight,2)+ pow(postfireWeightStdDev/postfireWeight,2)+pow(slopeError/slope, 2))

   return (dateYears,dateCalendar,ADBC,dateErrorYears)

def regression_figure(evenWeightPlot,xEvenArray, yEvenArray,yPercentEvenArray,weightSubPlot,xarray,yarray,xminval,xmaxval,ffilename,originalWeightAverage,originalWeightStdDev,prefireWeightAverage,prefireWeightStdDev,postfireWeightAverage,postfireWeightStdDev,weightChange, percentWeightChange, runDate,sampleName):
   #weightSubPlot.tick(labelsize=8)
   sx=array('d',[])
   sy=array('d',[])
   ex=array('d',[])
   ey=array('d',[])
   ppy=array('d',[])
   ppx=array('d',[])
   count=0
   scount=0
   originalWeight=0.0
   ecount=0
   secount=0

   for var in xarray:
      if var>xminval and var<xmaxval:
         sx.append(var)
         sy.append(yarray[count])
         scount+=1
      count += 1

   for var in xEvenArray:
      if var>xminval and var<xmaxval:
         ex.append(var)
         ey.append(yEvenArray[ecount])
         ppx.append(var)
         ppy.append(yPercentEvenArray[ecount])
         secount+=1
      ecount += 1

   minx=sx[1]
   maxx=sx[len(sx)-1]

   miny=sy[1]
   maxy=sy[len(sy)-1]

   minpy=ppy[1]
   maxpy=ppy[len(ppy)-1]

   sy=smoothListGaussian(sy)
   sx=smoothListGaussian(sx)
   spy=smoothListGaussian(ppy)

   #print xEvenArray
   slope, intercept, r_value, p_value, std_err = stats.linregress(sx,sy)
   results=""
   eslope,eintercept,er_value,ep_value,estd_err = stats.linregress(ex,ey)

   pslope,pintercept,pr_value,pp_value,pstd_err = stats.linregress(ex,ppy)

   linear_plot(evenWeightPlot,ex,ppy,ex[1],ex[len(ex)-1],ppy[1],ppy[len(ppy)-1],pslope,pintercept,pr_value,pp_value,ffilename)

   linear_plot(weightSubPlot,sx,sy,minx,maxx,miny,maxy,slope,intercept,p_value,r_value,ffilename)

   print "Original Weight: ", originalWeightAverage
   print "Weight Change: ", weightChange
   print "Percent Weight Change: ", percentWeightChange

   (dateYears,dateCalendar,ADBC,dateErrorYears)=age_calculate(slope,std_err,postfireWeightAverage,
      postfireWeightStdDev,prefireWeightAverage,prefireWeightStdDev)

   #(dateYears,dateCalendar,ADBC,dateErrorYears)=age_calculate(eslope,estd_err,postfireWeightAverage,
   #   postfireWeightStdDev,prefireWeightAverage,prefireWeightStdDev)

   (pdateYears,pdateCalendar,pADBC,pdateErrorYears)=age_calculateFromPercent(pslope,pstd_err,percentWeightChange)

   ws.cell('H10').value="Absolute Weight Slope"
   ws.cell('I10').value=eslope
   ws.cell('H11').value="Absolute WeightSlope StdErr"
   ws.cell('I11').value=estd_err
   ws.cell('H12').value="Absolute Weight Slope R-value"
   ws.cell('I12').value=er_value
   ws.cell('H13').value="Percent Weight Slope"
   ws.cell('I13').value=pslope
   ws.cell('H14').value="Age (BP)"
   ws.cell('I14').value=dateYears
   ws.cell('H15').value="Age (AD/BC)"
   ws.cell('I15').value=dateCalendar
   ws.cell('J16').value=ADBC
   ws.cell('H16').value="Error"
   ws.cell('I17').value=dateErrorYears
   ws.cell('H17').value="Age (BP) from %"
   ws.cell('I17').value=pdateYears
   ws.cell('H18').value="Age (AD/BC) from %"
   ws.cell('I18').value=pdateCalendar

   wb.save(filename = ffilename)

   results += "Results for sample: " + str(sampleName) + " from file: " +ffilename + "\n\r"
   results += "----------------------------------------------"+"\n\r"
   results += "Original Weight (g): " +str(originalWeightAverage)+" +/-" + str(originalWeightStdDev)+"\n\r"
   results += "Weight due to liquid H20:  "+ str(originalWeightAverage - prefireWeightAverage)+"\n\r"
   results += "Critical point (g): " + str(postfireWeightAverage+(originalWeightAverage-prefireWeightAverage))+"\r\n"
   results += "Prefire Absolute Weight (g): " + str(prefireWeightAverage)  + "+/-" + str(prefireWeightStdDev)+"\n\r"
   results += "Postfire Absolute Weight (g): " + str(postfireWeightAverage) + "+/-" + str(postfireWeightStdDev)+"\n\r"
   results += "----------------------------------------------"+"\n\r"
   results += "Difference in Absolute Weight (105 - 550):      " + str(prefireWeightAverage - postfireWeightAverage)+"\n\r"
   results += "RHX Slope (g/m^1/4):           " + str(eslope) +"\n\r"
   results += "RHX Intercept (g):       " + str(eintercept)+"\n\r"
   results += "RHX Slope r-value:         " + str(er_value)+"\n\r"
   results += "RHX Slope std_err:         " + str(estd_err)+"\n\r"
   results += "Absolute Weight Change Age (Years):     " + str(dateYears) + " +/- "  + str(dateErrorYears)+"\n\r"
   results += "Absolute Weight Change Date:            " + str(dateCalendar) + " " + str(ADBC) + " +/-" + str(dateErrorYears)+"\n\r"
   results += "----------------------------------------------"+"\n\r"
   results += "Percent Weight Change (105 - 550) (%):        " + str(percentWeightChange)+"\n\r"
   results += "Percent Weight Change RHX Slope (%/m^1/4):         " + str(pslope)+"\n\r"
   results += "Percent Weight Change RHX Intercept (g):       " + str(eintercept)+"\n\r"
   results += "Percent Weight Change RHX Slope r-value:         " + str(er_value)+"\n\r"
   results += "Percent Weight Change RHX Slope std_err:         " + str(estd_err)+"\n\r"
   results += "Percent Weight Change Age (Years): " + str(pdateYears) + " (BP) - " + str(pdateCalendar) + " " + pADBC + "\n\r"
   results += "Percent Weight Change Date:            " + str(pdateCalendar) + " " + str(pADBC) + " +/-" + str(pdateErrorYears)+"\n\r"
   results += "----------------------------------------------"+"\n\r"

   print results

   return results

def markEvenlySpacedPoints( x, y,ypercent,interval):
   number=1
   nextline=0
   x=smoothListGaussian(x)
   y=smoothListGaussian(y)
   ypercent=smoothListGaussian(ypercent)

   evenlySpacedXArray=array('d',[])
   evenlySpacedYArray=array('d',[])
   evenlySpacedYPercentArray=array('d',[])
   markcount=0
   skipcount=1
   for xval in x:
      if number == 1 or number==nextline:
         markcount += 1
         #print "number: ", number, "-", x[number],"-",y[number]
         evenlySpacedXArray.append(x[number])
         evenlySpacedYArray.append(y[number])
         evenlySpacedYPercentArray.append(ypercent[number])
         ## update sql table
         #try:
            #c=conn.cursor()
            #c.execute('update tblMeasurement set i_evenSpacePoint=1 where i_measurementID=%d' % row_list[number])
            #conn.commit()
         #except sqlite.OperationalError, msg:
            #logger.error( "A SQL error occurred: %s", msg)
         nextline=int(number+int(pow(skipcount,1.25)))
         skipcount +=1
         #print "X: ", xval, " Y:", y[number], " MARK"
      #print "X: ", xval, " Y:", y[number]
      number +=1
   return evenlySpacedXArray, evenlySpacedYArray,evenlySpacedYPercentArray

############################################################################################
############################################################################################


logger=logging.getLogger("rhxAnalyzeExcel")
logger.setLevel(logging.DEBUG)
# create file handler which logs even debug messages
fh = logging.FileHandler('analyzeExcel.log')
fh.setLevel(logging.DEBUG)
# create console handler with a higher log level
ch = logging.StreamHandler()
ch.setLevel(logging.ERROR)
# create formatter and add it to the handlers
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
fh.setFormatter(formatter)
ch.setFormatter(formatter)
# add the handlers to the logger
logger.addHandler(fh)
logger.addHandler(ch)

directory=''
if  os.name == 'nt':
   directory='c:\\Users\\Archy\\Dropbox\\Rehydroxylation\\'
elif os.name == 'mac':
   directory='/Users/clipo/Dropbox/Rehydroxylation/'

ffilename= easygui.fileopenbox(msg='Excel File', title='select file', filetypes=['*.xlsx'], default=directory)

wb = load_workbook(filename = ffilename )
listOfWorksheets=wb.get_sheet_names()
worksheet=ask_for_worksheet(listOfWorksheets)
ws = wb.get_sheet_by_name(name = worksheet)

sampleName =ws.cell('C5').value
operator = ws.cell('C3').value
dateOfRun=ws.cell('C15').value
notes=str(ws.cell('C7').value)+" "+str(ws.cell('C8').value)

## first check for weight values
originalWeighAverage=ws.cell('I1').value
if originalWeighAverage is None:
   originalWeightAverage,originalWeightStdDev,prefireWeightAverage,prefireWeightStdDev,postfireWeightAverage,postfireWeightStdDev = ask_for_weight_values()

   ws.cell('H1').value="Original Weight (g)"
   ws.cell('I1').value=originalWeightAverage
   ws.cell('H2').value="Original Weight StdDev)"
   ws.cell('I2').value=originalWeightStdDev
   ws.cell('H3').value="105 Degree C Weight (g)"
   ws.cell('I3').value=prefireWeightAverage
   ws.cell('H4').value="105 Degree C Weight StdDev"
   ws.cell('I4').value=prefireWeightStdDev
   ws.cell('H5').value="550 Degree C Weight (g)"
   ws.cell('I5').value=postfireWeightAverage
   ws.cell('H6').value="550 Degree C Weight StdDev"
   ws.cell('I6').value=postfireWeightStdDev

   weightChange=prefireWeightAverage-postfireWeightAverage
   percentWeightChange=weightChange/prefireWeightAverage
   ws.cell('H7').value="Weight Change (g)"
   ws.cell('I7').value=weightChange
   ws.cell('H8').value="Percent Weight Change (%)"
   ws.cell('I8').value=percentWeightChange

   ## delay since firing
   delay=float(ask_for_time_delay())
   if delay=="":
      delay=0.0
   ws.cell('H9').value="Delay Since Firing (minutes)"
   ws.cell('I9').value=delay

else:
   originalWeightAverage=float(ws.cell('I1').value)
   originalWeightStdDev=float(ws.cell('I2').value)
   prefireWeightAverage=float(ws.cell('I3').value)
   prefireWeightStdDev=float(ws.cell('I4').value)
   postfireWeightAverage=float(ws.cell('I5').value)
   postfireWeightStdDev=float(ws.cell('I6').value)
   weightChange=float(ws.cell('I7').value)
   percentWeightChange=float(ws.cell('I8').value)
   delay=float(ws.cell('I9').value)

#print "delay:  ", delay
numberOfSamples=1
wb.save(filename = ffilename)


## for each sample get the data for the post-fire
sampleNumber=1

while sampleNumber<=numberOfSamples:
   #humidity, coefficient = ask_for_RHCoefficient()
   params = {'figure.figsize': [8,8],'text.fontsize': 8, 'xtick.labelsize': 8,'ytick.labelsize': 8, 'axes.fontsize':8,'axes.style' : 'plain', 'axes.axis':'both'}
   plt.rcParams.update(params)
   fig=plt.figure(sampleNumber)
   fig.suptitle(sampleName, fontsize=10)

   ## define all of the plots in the figure
   ax1 = fig.add_subplot(4,2,1)
   weightSubPlot = fig.add_subplot(4,2,2)
   #filteredWeightPlot=fig.add_subplot(6,2,3)
   tempPlot = fig.add_subplot(4,2,3)               # filtered weight graph
   humidPlot = fig.add_subplot(4,2,4)              # the humidity graph
   slopeGraph = fig.add_subplot(4,2,5)
   dateGraph = fig.add_subplot(4,2,6)
   evenPlot = fig.add_subplot(4,2,7)
   evenPlotRegression = fig.add_subplot(4,2,8)

   finish=0

   preFireWeight=0.0
   postFireWeight=0.0
   print "Sample Name:  ", sampleName
   print "Operator:  ", operator
   print "Run Date: ", dateOfRun
   print "Notes:  ",notes
   runDate = datetime.strptime(str(dateOfRun), '%m-%d-%Y')

   while (finish==0):

      count=0
      x=array('d',[])
      y=array('d',[])
      ypercent=array('d',[])
      temp=array('d',[])
      humid=array('d',[])
      tempSmooth=array('d',[])
      humidSmooth=array('d',[])
      y_correct=array('d',[])
      row_list=array('I',[])
      humidAverage=0.0
      humidStdDev=0.0
      tempAverage=0.0
      tempStdDev=0.0
      rowNumber=20
      endOfData=0
      initialWeight=0.0

      while endOfData==0:
         count=0
         xv=ws.cell('A'+str(rowNumber)).value
         #print xv
         if xv is None:
            break
         xv = float(xv) + delay
         #print xv
         x.append(pow(float(xv),0.25))
         y.append((ws.cell('B'+str(rowNumber)).value)/1000) # turn to grams
         ypercent.append(float(ws.cell('C'+str(rowNumber)).value))
         temp.append(ws.cell('D'+str(rowNumber)).value)
         humid.append(ws.cell('E'+str(rowNumber)).value)
         if rowNumber ==20:
            intitalWeight= float(ws.cell('C'+str(rowNumber)).value)

         rowNumber +=1

      evenlySpacedYPercentArray=array('d',[])
      evenlySpacedXArray=array('d',[])
      evenlySpacedYArray=array('d',[])
      ## now create an array that has the evenly spaced values (and update the sql table with these values)
      interval=10
      liquidwater=postfireWeightAverage+(originalWeightAverage-prefireWeightAverage)

      evenlySpacedXArray, evenlySpacedYArray, evenlySpacedYPercentArray= markEvenlySpacedPoints(x, y,ypercent,interval)

      evenlySpacedPlot(evenPlot,evenlySpacedXArray,evenlySpacedYPercentArray,liquidwater)

      overall_plot(ax1,x,y,sampleNumber,liquidwater)

      temp_humidity_plot(tempPlot,humidPlot,x,temp,humid)

      tempSmooth = smoothListGaussian(temp)
      humidSmooth = smoothListGaussian(humid)

      fig2=plt.figure(sampleNumber)
      fig2.suptitle("SLOPE", fontsize=10)

      slopeArray=zeros(alen(x))
      interceptArray=zeros(alen(x))
      r_valueArray=zeros(alen(x))
      p_valueArray=zeros(alen(x))
      std_errArray=zeros(alen(x))
      dateArray=zeros(alen(x))

      percentSlopeArray=zeros(alen(x))

      ### size of the window to calculate slope, intercept, etc. ####
      window = 3
      #########################################

      (slopeArray,interceptArray,r_valueArray,p_valueArray,std_errArray,dateArray)=graph_slope(
         sampleNumber,evenlySpacedXArray,evenlySpacedYPercentArray,window,originalWeightAverage,originalWeightStdDev,
         prefireWeightAverage,prefireWeightStdDev,postfireWeightAverage,postfireWeightStdDev)
      slopeGraph.plot(evenlySpacedXArray, slopeArray, '.')
      slopeGraph.set_xlabel('Time (m^1/4)', fontsize=8)
      slopeGraph.set_ylabel('Slope', fontsize=8)
      slopeGraph.ticklabel_format(style='plain', axis='both, fontsize=8')
      slopeGraph.grid(True)

      dateGraph.set_xlabel('Time (m^1/4)', fontsize=8)
      dateGraph.set_ylabel('Date (Years)', fontsize=8)
      dateGraph.ticklabel_format(style='plain', axis='both, fontsize=8')
      dateGraph.grid(True)

      dateGraph.plot(evenlySpacedXArray,dateArray,'--')
      #plt.plot(xArray,ainterceptArray,'x')
      plt.ion()     # turns on interactive mode
      plt.show()    # now this should be non-blocking\

      minval=""
      maxval=""
      results=""

      minval, maxval=ask_for_minmax_values(min(x),max(x))

      results += "NON FILTERED RESULTS"+"\n\r"
      results += regression_figure(evenPlotRegression,
         evenlySpacedXArray,evenlySpacedYArray,evenlySpacedYPercentArray,
         weightSubPlot,
         x,y,minval,maxval,
         ffilename,originalWeightAverage,originalWeightStdDev,prefireWeightAverage,prefireWeightStdDev,postfireWeightAverage,postfireWeightStdDev,weightChange, percentWeightChange,runDate,sampleName)



      easygui.codebox(msg=None, title='Results ', text=results)

      test=""

      msg     = "What would you like to do?"
      choices = ["Try Again","Quit"]
      reply   = easygui.buttonbox(msg,image=None,choices=choices)

      if reply=="Try Again":
         plt.clf()
         break
      else:
         exit()





