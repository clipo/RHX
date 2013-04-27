__author__ = 'clipo'

import sys
import logging
import os
import sqlite3 as sqlite
from datetime import date
from datetime import datetime

import numpy as np
from scipy import stats
from lmfit import minimize, Parameters, Parameter, report_errors
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl import Workbook
from scipy import *
#from scipy.optimize import leastsq
#import scikits.datasmooth as ds
#from lmfit import minimize, Parameters, Parameter, report_errors
import easygui

###CONSTANTS
CALCULATE_AGE_FROM_PERCENTS = False
DEBUG = 0

#import rpy2.robjects as robjects
#from rpy2.robjects import FloatVector
#from rpy2.robjects.packages import importr

#stats = importr('stats')
#base = importr('base')
#minpack=importr('minpack.lm')

# define objective function: returns the array to be minimized
def fcn2min(params, x, data):
    """ model decaying sine wave, subtract data"""
    alpha = params['alpha'].value
    power = params['power'].value
    #model = pow((data/alpha),power)
    #return model - x
    model = alpha * pow(x, power)
    return model - data


def quarterPowerFunction(params, x, data):
    alpha = params['alpha'].value
    model = alpha * pow(x, 0.25)
    #model = pow((x/alpha),4)
    return model - data


def smoothTriangle(data, degree, dropVals=False):
    """performs moving triangle smoothing with a variable degree."""
    """note that if dropVals is False, output length will be identical
    to input length, but with copies of data at the flanking regions"""
    triangle = np.array(range(degree) + [degree] + range(degree)[::-1]) + 1
    smoothed = []
    for i in range(degree, len(data) - degree * 2):
        point = data[i:i + len(triangle)] * triangle
        smoothed.append(sum(point) / sum(triangle))
    if dropVals: return smoothed
    smoothed = [smoothed[0]] * (degree + degree / 2) + smoothed
    while len(smoothed) < len(data): smoothed.append(smoothed[-1])
    return smoothed


def smooth(y, window_len=11, window='hanning'):
    """smooth the data using a window with requested size.

    This method is based on the convolution of a scaled window with the signal.
    The signal is prepared by introducing reflected copies of the signal
    (with the window size) in both ends so that transient parts are minimized
    in the begining and end part of the output signal.

    input:
        x: the input signal
        window_len: the dimension of the smoothing window; should be an odd integer
        window: the type of window from 'flat', 'hanning', 'hamming', 'bartlett', 'blackman'
            flat window will produce a moving average smoothing.

    output:
        the smoothed signal

    example:

    t=linspace(-2,2,0.1)
    x=sin(t)+randn(len(t))*0.1
    y=smooth(x)

    see also:

    numpy.hanning, numpy.hamming, numpy.bartlett, numpy.blackman, numpy.convolve
    scipy.signal.lfilter

    TODO: the window parameter could be the window itself if an array instead of a string
    NOTE: length(output) != length(input), to correct this: return y[(window_len/2-1):-(window_len/2)] instead of just y.
    """

    if not window in ['flat', 'hanning', 'hamming', 'bartlett', 'blackman']:
        raise ValueError, "Window is on of 'flat', 'hanning', 'hamming', 'bartlett', 'blackman'"

    s = np.r_[y[window_len - 1:0:-1], y, y[-1:-window_len:-1]]
    #print(len(s))
    if window == 'flat': #moving average
        w = np.ones(window_len, 'd')
    else:
        w = eval('np.' + window + '(window_len)')

    y = np.convolve(w / w.sum(), s, mode='valid')

    return y[(window_len / 2 - 1):-(window_len / 2)]


def rhxFunction(a, x, p):
    return a * np.power(x, p)


def errorFunction(a, x, p, nly):
    return nly - rhxFunction(a, x, p)


def nlinRegression(sx, sy, minval, maxval):
    a_guess = 0.000
    kd, cov, infodict, mesg, ier = leastsq(errorFunction, guess=a_guess, args=(sx, sy), maxfev=10000, full_output=True)
    print "nonlinear regression message: ", mesg
    print "nonlinear regression code: ", ier
    return kd[0]


def dateCalcFromNlinRegression(alpha, weightChange):
    timeElapsed = pow((weightChange / alpha), 4)
    return timeElapsed


def minToYears(dateMinutes):
    mdate = date.today()
    dateHours = dateMinutes / 60
    ## 24 hours to day
    dateDays = dateHours / 24
    ## 365 hours/year
    dateYears = dateDays / 365.25
    ## AD/BC - subtract from this year
    ## initialize the maxDate variable with a date (today)
    ADBC = ""
    maxDateYear = mdate.year
    ## now measure the difference-- the year at the end of the measurement minus number of years in calculations
    dateCalendar = maxDateYear - dateYears
    if dateCalendar <= 0:
        ADBC = "BC"
    else:
        ADBC = "AD"

    return dateYears, dateCalendar, ADBC


def ask_for_excel_or_sqlite():
    msg = "Choose analysis type:"
    choices = ["VTI-SA Data", "RHX Environmental Chamber", "Export Data from Sqlite", "Analyze Aggregate Values",
               "Update Mean Values", "Quit"]
    reply = easygui.buttonbox(msg, choices=choices)
    if reply == "VTI-SA Data":
        return 1
    elif reply == "RHX Environmental Chamber":
        return 2
    elif reply == "Export Data from Sqlite":
        return 3
    elif reply == "Analyze Aggregate Values":
        return 4
    elif reply == "Update Mean Values":
        return 5
    else:
        return 0


def smoothListGaussian(xlist, ylist, strippedXs=False, degree=5):
    window = degree * 2 - 1
    weight = np.array([1.0] * window)
    weightGauss = zeros(window)
    count = 0
    for i in range(window):
        i = i - degree + 1
        frac = i / float(window)
        gauss = 1 / (np.exp((4 * frac) ** 2))
        weightGauss[count] = gauss
        count += 1
    weight = np.array(weightGauss) * weight
    #smoothedY = [0.0] * (len(ylist) - window)
    #print "length of ylist: ", alen(ylist)
    #print "length of window: ", window
    smoothedY = np.zeros((alen(ylist) - window))
    #smoothedX = [0.0] * (len(xlist) - window)
    smoothedX = np.zeros((alen(xlist) - window))
    for i in range(smoothedY.size):
        smoothedY[i] = sum(np.array(xlist[i:i + window]) * weight) / sum(weight)
        smoothedX[i] = xlist[i]
    return smoothedX, smoothedY


def evenlySpacedPlot(evenPlot, evenXarray, evenYarray):
    evenPlot.scatter(evenXarray, evenYarray, marker='o', c='r')
    evenPlot.set_xlabel('Time (m^1/4)', fontsize=8)
    evenPlot.set_ylabel('Mass change (% of initial mass)', fontsize=8)
    ymin = min(evenYarray)
    ymax = max(evenYarray)
    evenPlot.set_ylim(ymin, ymax)
    evenPlot.grid(True)
    plt.ion()     # turns on interactive mode
    plt.show()    # now this should be non-blocking


def temp_humidity_plot(tempPlot, humidPlot, timeArray, tempArray, humidArray):
    #tempPlot.tick(labelsize=8)
    #humidPlot.tick(labelsize=8)
    tempPlot.scatter(timeArray, tempArray, marker='^', c='r')
    humidPlot.scatter(timeArray, humidArray, marker='o', c='g')
    tempPlot.set_xlabel('Time (m^1/4)', fontsize=8)
    humidPlot.set_xlabel('Time (m^1/4)', fontsize=8)
    tempPlot.set_ylabel('Non-Filtered Temperature (C)', fontsize=8)
    humidPlot.set_ylabel('Non-Filtered Humidity (%RH)', fontsize=8)
    tempPlot.grid(True)
    humidPlot.grid(True)
    plt.ion()     # turns on interactive mode
    plt.show()    # now this should be non-blocking


def correctForRH( xArray, yArray, humidArray, humiditySetting, rhCoefficent ):
    correctedArray = zeros(alen(yArray))
    num = 0
    for oldWeight in yArray:
        humidity = humidArray[num]
        newWeight = oldWeight + rhCoefficent * (humidity - humiditySetting) / humiditySetting
        correctedArray[num] = newWeight
        num += 1
    return correctedArray

## if the sqlite file is already in existence and you have the file handle -- just need to do this...
def reopenDatabase():
    dbfilename = easygui.fileopenbox(msg='Open an existing sqlite file for this set of samples.', title='Open Database',
                                     default="C:/Users/Archy/Dropbox/Rehydroxylation/Logger/Data/*.sqlite",
                                     filetypes='*.sqlite')
    #print "filename:", filename
    if dbfilename is None:
        return
    global conn
    try:
        conn = sqlite.connect(dbfilename, detect_types=sqlite.PARSE_DECLTYPES, isolation_level=None)
        path = '../../RequiredFiles/libsqlitefunctions.so'
        # open the sqlite database and create a connection
        # create a cursor to the database
        conn.enable_load_extension(True)
        conn.load_extension(path)
        return True
    except sqlite.OperationalError, msg:
        logger.error("A SQL error occurred trying to open sqlitefile with functions: %s", msg)
        return False


def alertWindows(message):
    easygui.msgbox(msg=message, title="Warning")
    return


def getCurrentStateOfRun():
    c = conn.cursor()
    try:
        c.execute('select max(i_preOrPost) as max from tblMeasurement')
        row = c.fetchone()
        if row[0] is None:
            return -1
        return int(row[0])
    except sqlite.OperationalError, msg:
        logger.error("A SQL error occurred: %s", msg)
        return -1

## Get the average weight and std dev of weight for this crucible.
def getCrucibleWeight(runID, positionNumber):
    try:
        c = conn.cursor()
        t = (runID, positionNumber)
        #print "RUNID:  ",runID, "PositionNumber:",positionNumber
        averageWeight = float(0.0)
        c.execute(
            'select f_averageWeight, f_stdevWeight from tblCrucible where i_runID=? and i_positionNumber=?', t)
        logger.debug(
            'select f_averageWeight, f_stdevWeight from tblCrucible where i_runID=%d and i_positionNumber=%d' % t)
        data = c.fetchone()
        if data is None:
            return False
        else:
            #print "sample: ", positionNumber, " Crucible Weight: ", data[0],"+/-", data[1]
            return float(data[0]), float(data[1])
    except sqlite.OperationalError, msg:
        logger.error("A SQL error occurred trying to getCrucibleWeight: %s", msg)
        return False


def updateAndFixAverageValues():
    value = reopenDatabase()
    if value is False:
        alertWindows("Database open failed!")
        return
        # first get the number of samples
    v_operatorName, v_locationCode, i_numberOfSamples, t_assemblageName, f_locationTemperature, humidityLocation, d_dateTimeFiring, i_durationOfFiring, notes, status = getRunInfo(
        1)
    currentMaxState = getCurrentStateOfRun()
    currentMaxState += 1
    stateCounter = 0
    while stateCounter < currentMaxState:
        sampleCounter = 1
        runID = 1
        while sampleCounter <= i_numberOfSamples:
            ### first check to see if there are data there at all for whatever state we are in
            measurementForState = 0
            try:
                t = (sampleCounter, stateCounter)
                c = conn.cursor()
                c.execute("select count(*) from tblMeasurement where i_positionNumber=? and i_preOrPost =?", t)
                row = c.fetchone()
                measurementForState = row[0]
                #print "For sample: ", sampleCounter, " State: ", stateCounter, " Number of measurements: ", measurementForState
            except sqlite.OperationalError, msg:
                logger.error("A SQL error occurred on getting count of tblMeasurement: %s", msg)

            if measurementForState > 0:
                #print "State:  ", stateCounter, "  Sample # ", sampleCounter
                crucibleWeightAverage, crucibleWeightStdDev = getCrucibleWeight(1, sampleCounter)

                try:
                    c = conn.cursor()
                    t = ( sampleCounter, stateCounter )
                    print(
                        'select avg(f_weight) as AVE, stdev(f_weight) as SD from tblMeasurement where i_positionNumber= %d '
                        'and i_preOrPost= %d and i_runID=1' % t)
                    c.execute(
                        'select avg(f_weight) as AVE, stdev(f_weight) as SD from tblMeasurement where i_positionNumber=? and i_preOrPost=? and i_runID=1',
                        t)
                    data = c.fetchone()
                    averageWeight = float(data[0])
                    stddevWeight = float(data[1])
                    #print "Average Measurement Weight: ", averageWeight, " +/- ", stddevWeight
                except sqlite.OperationalError, msg:
                    logger.error("A SQL error occurred when I tried to select averages and std dev: %s", msg)
                    return False
                sampleWeight = float(averageWeight - crucibleWeightAverage)
                #print "Average Sherd Weight: ", sampleWeight, " +/-", stddevWeight

                if stateCounter == 0:
                    try:
                        t = ( averageWeight, stddevWeight, averageWeight, stddevWeight, sampleCounter )
                        print(
                            'update tblCrucible set f_averageWeight=%f, f_stdevWeight=%f, f_emptyWeightAverage=%f, f_emptyWeightStdDev=%f where i_runID=1 and i_positionNumber=%d  ' % t)
                        c.execute(
                            'update tblCrucible set f_averageWeight=?, f_stdevWeight=?, f_emptyWeightAverage=?, f_emptyWeightStdDev=? where i_runID=1 and i_positionNumber=?  ',
                            t)
                        conn.commit()
                        t = ( averageWeight, stddevWeight, sampleCounter )
                        print(
                            'update tblSample set f_crucibleWeightAverage=%f, f_crucibleWeightStdDev=%f where i_runID=1 and i_positionNumber=%d  ' % t)
                        c.execute(
                            'update tblSample set f_crucibleWeightAverage=?, f_crucibleWeightStdDev=? where i_runID=1 and i_positionNumber=?  ',
                            t)
                        conn.commit()
                        c.execute(
                            'select  f_averageWeight, f_stdevWeight from tblCrucible where  i_runID=1 and i_positionNumber=%d  ' % sampleCounter)
                        data = c.fetchone()
                        #print "Values -- averageWeight: %s  stdDev: %s ", (data['f_averageWeight'],data['f_stdevWeight'])
                    except sqlite.OperationalError, msg:
                        logger.error("A SQL error occurred on the update tblCrucible (state 1): %s", msg)
                        return False

                elif stateCounter == 1:
                    try:
                        t = ( sampleWeight, stddevWeight, sampleCounter )
                        print(
                            'update tblSample set f_sherdWeightInitialAverage=%f, f_sherdWeightInitialStdDev=%f where i_runID=1 and i_positionNumber=%d  ' % t)
                        c.execute(
                            'update tblSample set f_sherdWeightInitialAverage=?, f_sherdWeightInitialStdDev=? where i_runID=1 and i_positionNumber=? ',
                            t)
                        conn.commit()
                        c.execute(
                            'select  f_sherdWeightInitialAverage, f_sherdWeightInitialStdDev from tblSample where  i_runID=1 and i_positionNumber=%d  ' % sampleCounter)
                        data = c.fetchone()
                        #print "Values -- averageWeight: %s  stdDev: %s ", (data['f_sherdWeightInitialAverage'],data['f_sherdWeightInitialStdDev'])
                    except sqlite.OperationalError, msg:
                        logger.error("A SQL error occurred on the update tblSample (state 1): %s", msg)
                        return False

                elif stateCounter == 2:
                    try:
                        t = ( averageWeight, stddevWeight, sampleCounter )
                        print(
                            'update tblCrucible set f_105WeightAverage=%f, f_105WeightStdDev=%f  where i_runID=1 and i_positionNumber=%d  ' % t)
                        c.execute(
                            'update tblCrucible set f_105WeightAverage=?, f_105WeightStdDev=? where i_runID=1 and i_positionNumber=?  ',
                            t)
                        conn.commit()
                        c.execute(
                            'select  f_105WeightAverage, f_105WeightStdDev from tblCrucible where  i_runID=1 and i_positionNumber=%d  ' % sampleCounter)
                        data = c.fetchone()
                        #print "Values -- averageWeight: %s  stdDev: %s ", (data['f_105WeightAverage'],data['f_105WeightStdDev'])
                    except sqlite.OperationalError, msg:
                        logger.error("A SQL error occurred on the update tblCrucible (state 2): %s", msg)
                        return False
                    try:
                        t = ( sampleWeight, stddevWeight, sampleCounter )
                        print(
                            'update tblSample set f_preWeightAverage=%f, f_preWeightStdDev=%f where i_runID=1 and i_positionNumber=%d  ' % t)
                        c.execute(
                            'update tblSample set f_preWeightAverage=?, f_preWeightStdDev=? where i_runID=1 and i_positionNumber=?  ',
                            t)
                        conn.commit()
                    except sqlite.OperationalError, msg:
                        logger.error("A SQL error occurred in the update tblSample (state 2): %s", msg)
                        return False

                elif stateCounter == 3:
                    try:
                        t = ( averageWeight, stddevWeight, sampleCounter )
                        print(
                            'update tblCrucible set f_550WeightAverage=%f, f_550WeightStdDev=%f where i_runID=1 and i_positionNumber=%d  ' % t)
                        c.execute(
                            'update tblCrucible set f_550WeightAverage=?, f_550WeightStdDev=? where i_runID=1 and i_positionNumber=?  ',
                            t)
                        conn.commit()
                        c.execute(
                            'select  f_550WeightAverage, f_550WeightStdDev from tblCrucible where  i_runID=1 and i_positionNumber=%d  ' % sampleCounter)
                        data = c.fetchone()
                        #print "Values -- averageWeight: %s  stdDev: %s ", (data['f_550WeightAverage'],data['f_550WeightStdDev'])
                    except sqlite.OperationalError, msg:
                        logger.error("A SQL error occurred on the update tblCrucible (state 3): %s", msg)
                        return False

                    try:
                        t = ( sampleWeight, stddevWeight, sampleCounter )
                        print(
                            'update tblSample set f_postFireWeightAverage=%f, f_postFireWeightStdDev=%f where i_runID=1 and i_positionNumber=%d  ' % t)
                        c.execute(
                            'update tblSample set f_postFireWeightAverage=?, f_postFireWeightStdDev=? where i_runID=1 and i_positionNumber=?  ',
                            t)
                        conn.commit()
                        c.execute(
                            'select  f_postFireWeightAverage, f_postFireWeightStdDev from tblSample where  i_runID=1 and i_positionNumber=%d  ' % sampleCounter)
                        data = c.fetchone()
                        #print "Values -- averageWeight: %s  stdDev: %s ", (data['f_postFireWeightAverage'],data['f_postFireWeightStdDev'])
                    except sqlite.OperationalError, msg:
                        logger.error("A SQL error occurred on the update tblSample (state 3): %s", msg)
                        return False

            sampleCounter += 1
        stateCounter += 1
    closeDatabase()
    msg = "Data update is complete."
    title = "Done"
    if easygui.ccbox(msg, title):     # show a Continue/Cancel dialog
        return "continue"
    else:
        return "cancel"


def filtered_temp_humidity_plot(filteredTempPlot, filteredHumidPlot, timeArray, filteredTempArray, filteredHumidArray):
    #filteredTempPlot.tick(labelsize=8)
    filteredTempPlot.scatter(timeArray, filteredTempArray, marker='^', c='r')
    filteredHumidPlot.scatter(timeArray, filteredHumidArray, marker='o', c='g')
    filteredTempPlot.set_xlabel('Time (m^1/4)', fontsize=8)
    filteredHumidPlot.set_xlabel('Time (m^1/4)', fontsize=8)
    filteredTempPlot.set_ylabel('Filtered Temperature (C)', fontsize=8)
    filteredHumidPlot.set_ylabel('Filtered Humidity (%RH)', fontsize=8)
    plt.ion()     # turns on interactive mode
    plt.show()    # now this should be non-blocking


def overall_plot(ax1, xArray, yArray, sampleNumber):
    #ax1.tick(labelsize=8)
    ymin = np.amin(yArray)
    ymax = np.amax(yArray)
    ax1.scatter(xArray, yArray, marker='x', c='r')
    ax1.set_ylim(ymin, ymax)
    ax1.set_xlabel('Time (m^1/4)', fontsize=8)
    ax1.set_ylabel('Weight (g)', fontsize=8)
    ax1.grid(True)
    plt.ion()     # turns on interactive mode
    plt.show()    # now this should be non-blocking


def ask_for_RHCoefficient():
    msg = "Enter Humidity Information"
    title = "Humidity"
    fieldNames = ["RH Setting (%RH)", "Coefficient for Weight Correlation to RH"]
    fieldValues = []  # we start with blanks for the values
    fieldValues = easygui.multenterbox(msg, title, fieldNames)
    # make sure that none of the fields was left blank
    while 1:  # do forever, until we find acceptable values and break out
        if fieldValues is None:
            break
        errmsg = ""
        if fieldValues[0] < 1 or fieldValues[1] < 1:
            errmsg += "You must enter a postive number."
            # look for errors in the returned values
        for i in range(alen(fieldNames)):
            if fieldValues[i].strip() == "":
                errmsg += '"%s" is a required field.\n\n' % fieldNames[i]
        if errmsg == "":
            break # no problems found
        else:
            # show the box again, with the errmsg as the message
            errmsg += " Invalid values. Please try again. "
            fieldValues = easygui.multenterbox(errmsg, title, fieldNames, fieldValues)

    return float(fieldValues[0]), float(fieldValues[1])


def ask_for_time_delay():
    msg = "Enter time (minutes) since 500C firing ended"

    fieldValue = 0  # we start with blanks for the values
    fieldValue = easygui.enterbox(msg, "0")
    return fieldValue


def ask_for_worksheet(choices):
    if len(choices)==1:
        return choices[0]
    else:
        msg = "Choose the worksheet that contains the run data."
        title = "Worksheet Choice"
        choice = easygui.choicebox(msg, title, choices)
        return choice

def ask_for_weight_values():
    msg = "Enter Weight Data"
    title = "Weights"
    fieldNames = ["Initial Weight (g)", "Std Dev", "105 Degrees C Weight (g)", "Std Dev", "500 Degrees C Weight",
                  "Std Dev"]
    fieldValues = []  # we start with blanks for the values
    fieldValues = easygui.multenterbox(msg, title, fieldNames)
    # make sure that none of the fields was left blank
    while 1:  # do forever, until we find acceptable values and break out
        if fieldValues is None:
            break
        errmsg = ""
        if float(fieldValues[0]) < float(fieldValues[2]):
            errmsg += "The initial weight must be higher than the 105 degree weight "
            # look for errors in the returned values
        if float(fieldValues[2]) < float(fieldValues[4]):
            errmsg += "The 105 degree weight must be higher than the 500 degree weight "
            # look for errors in the returned values
        for i in range(len(fieldNames)):
            if fieldValues[i].strip() == "":
                errmsg += ('"%s" is a required field.\n\n' % fieldNames[i])
        if errmsg == "":
            break # no problems found
        else:
            # show the box again, with the errmsg as the message
            errmsg += " Invalid values. Please try again. "
            fieldValues = easygui.multenterbox(errmsg, title, fieldNames, fieldValues)

    return float(fieldValues[0]), float(fieldValues[1]), float(fieldValues[2]), float(fieldValues[3]), float(
        fieldValues[4]), float(fieldValues[5])


def ask_for_minmax_values(minx, maxx, errormessage):
    msg = errormessage + " Information for linear calculations"
    title = "Linear Calculations"
    fieldNames = ["Minimum time (m^1/4)", "Maximum time (m^1/4)", "Power (e.g., 4)"]
    values = [minx, maxx,4]  # we start with blanks for the values
    fieldValues = easygui.multenterbox(msg, title, fieldNames, values)
    # make sure that none of the fields was left blank
    while 1:  # do forever, until we find acceptable values and break out
        if fieldValues is None:
            break
        errmsg = ""
        if float(fieldValues[0]) > float(fieldValues[1]):
            errmsg += "The minimum value must be lower than the maximum value. "
            # look for errors in the returned values
        for i in range(len(fieldNames)):
            if fieldValues[i].strip() == "":
                errmsg += ('"%s" is a required field.\n\n' % fieldNames[i])
        if errmsg == "":
            break # no problems found
        else:
            # show the box again, with the errmsg as the message
            errmsg += " Invalid values. Please try again. "
            fieldValues = easygui.multenterbox(errmsg, title, fieldNames, fieldValues)
    if fieldValues is None:
        closeDatabase()
        sys.exit()
    return float(fieldValues[0]), float(fieldValues[1]), float(fieldValues[2])


def graph_slope(sampleNumber, x, y, window, prefireWeightAverage, prefireWeightStdDev, postfireWeightAverage,
                postfireWeightStdDev):
    xs = zeros(alen(x))
    ys = zeros(alen(x))
    slopeArray = zeros(alen(x))
    interceptArray = zeros(alen(x))
    r_valueArray = zeros(alen(x))
    p_valueArray = zeros(alen(x))
    std_errArray = zeros(alen(x))
    date_Array = zeros(alen(x))
    slope = 0.0
    intercept = 0.0
    r_value = 0.0
    p_value = 0.0
    std_err = 0.0
    date = 0.0
    num = 1
    dateYears = 0.0
    dateCalendar = 0.0
    ADBC = ""
    dateErrorYear = 0.0

    while num < alen(x):
        #print "Num: ", num
        if num < window:
            slopeArray[num] = 0.0
            interceptArray[num] = 0.0
            r_valueArray[num] = 0.0
            p_valueArray[num] = 0.0
            std_errArray[num] = 0.0
            #print "slope: 0"
        else:
            xs = x[num - window:num:1]
            ys = y[num - window:num:1]
            #print xs
            slope, intercept, r_value, p_value, std_err = stats.linregress(xs, ys)
            #print "slope:", slope, " -- ", num

            slopeArray[num] = slope
            interceptArray[num] = intercept
            r_valueArray[num] = r_value
            p_valueArray[num] = p_value
            std_errArray[num] = std_err
            slopeFourthPower = pow(float(slope), 4)
            power = 4
            dateYears, dateCalendar, ADBC, dateErrorYear, slopeFourthPower = age_calculate(power, slope, std_err,
                                                                                           intercept,
                                                                                           postfireWeightAverage,
                                                                                           postfireWeightStdDev,
                                                                                           prefireWeightAverage,
                                                                                           prefireWeightStdDev)
            date_Array[num] = dateYears
            #print dateYears
        num += 1
    return slopeArray, interceptArray, r_valueArray, p_valueArray, std_errArray, date_Array


def age_calculateFromPercent(power, pslope, pstd_err, percentWeightChange):
    if power is None:
        power = 4
        ## slope is grams/min^0.25
    ## convert to gram/min
    slopeFourthPower = pow(abs(float(pslope)), power)

    ## now divide weight by the slope

    dateMinutes = percentWeightChange / slopeFourthPower
    ## minutes to hours to days to years
    denominator = 365.25 * 60 * 24
    lower_dateSlope = pow((pslope - pstd_err), power)
    try:
        lower_dateMinutes = percentWeightChange / lower_dateSlope
    except ZeroDivisionError:
        lower_dateMinutes = 0.0
    lower_dateYears = lower_dateMinutes / denominator
    upper_dateSlope = pow((pslope + pstd_err), power)
    try:
        upper_dateMinutes = percentWeightChange / upper_dateSlope
    except ZeroDivisionError:
        upper_dateMinutes = 0.0
    upper_dateYears = upper_dateMinutes / denominator
    slope_errorYears = upper_dateYears - lower_dateYears

    dateYears, dateCalendar, ADBC = minToYears(dateMinutes)

    # now calculate error terms
    # this consists of the prefire weight error, the postfire weight error, the slope error,
    # where prefire-postfire/slope
    # so date * sqrt( ( prefireError/prefire)^2 + (postfireError/postfire)^2 + *slopeError/slope)^2)
    total_error = slope_errorYears

    return dateYears, dateCalendar, ADBC, total_error, slopeFourthPower


def filter_by_humidity( xArray, yArray, rhArray ):
    meanRH = np.average(rhArray)
    outX = np.array([])
    outY = np.array([])
    for index, value in enumerate(xArray):
        if (meanRH - 1) < rhArray[index] < (meanRH + 1):
            outX = np.append(outX, value)
            outY = np.append(outY, yArray[index])
    return outX, outY


def age_calculate(power, slope, slopeError, intercept, postfireWeight, postfireWeightStdDev, prefireWeight, prefireWeightStdDev):
    ## difference is the pre fire weight  minus intercept (post fire weight)
    if power is None:
        power = 4
    #diffWeight = prefireWeight - postfireWeight
    ## here we will use the postfireWeight as the intercept
    #intercept = postfireWeight  ## NOTE THAT IM RESETTING INTERCEPT TO POSTFIRE WEIGHT
    diffWeight = prefireWeight - postfireWeight

    ## slope is grams/min^0.25
    ## convert to gram/min
    slopeFourthPower = pow(abs(float(slope)), power)

    ## now divide weight by the slope and take to the 4th power
    dateMinutes = pow((diffWeight/ slope), power)
    dateYears, dateCalendar, ADBC = minToYears(dateMinutes)

    if postfireWeightStdDev is None:
        postfireWeightStdDev = 0.0

    ## minutes to hours to days to years
    denominator = 365.25 * 60 * 24

    lower_slopeDiff = slope - slopeError
    try:
        lower_dateMinutes = pow((diffWeight / lower_slopeDiff),power)
    except ZeroDivisionError:
        lower_dateMinutes = 0.0
    lower_dateYears = lower_dateMinutes / denominator

    upper_slopeDiff = slope + slopeError
    try:
        upper_dateMinutes = pow((diffWeight/ upper_slopeDiff),power)
    except ZeroDivisionError:
        upper_dateMinutes = 0.0
    upper_dateYears = upper_dateMinutes / denominator


    slope_errorYears = (upper_dateYears - lower_dateYears) / 2
    #print "Max years due to slope calculation error: ", slope_errorYears

    lower_diffWeight = diffWeight - (abs(postfireWeightStdDev) + abs(prefireWeightStdDev))
    try:
        lower_diffWeightMinutes = pow((lower_diffWeight / slope),power)
    except ZeroDivisionError:
        lower_diffWeightMinutes = 0.0
    lower_diffWeightYears = lower_diffWeightMinutes / denominator

    upper_diffWeight = diffWeight + (abs(postfireWeightStdDev) + abs(prefireWeightStdDev))
    try:
        upper_diffWeightMinutes = pow((upper_diffWeight / slope),power)
    except ZeroDivisionError:
        upper_diffWeightMinutes = 0.0
    upper_diffWeightYears = upper_diffWeightMinutes / denominator

    diffWeight_errorYears = (upper_diffWeightYears - lower_diffWeightYears) / 2
    #print "max years error due to weighing measurement error: ", diffWeight_errorYears

    # now calculate error terms
    # this consists of the prefire weight error, the postfire weight error, the slope error,
    # where prefire-postfire/slope
    # so date * sqrt( ( prefireError/prefire)^2 + (postfireError/postfire)^2 + *slopeError/slope)^2)
    ## just the sum of error?
    total_error = pow((pow(slope_errorYears,2) + pow(diffWeight_errorYears,2)),0.5)

    return dateYears, dateCalendar, ADBC, total_error, slopeFourthPower


def age_calculate_nlin(power, power_stderr, alpha, alpha_stderr, postfireWeight, postfireWeightStdDev, prefireWeight,
                       prefireWeightStdDev):
    ## difference is the pre fire weight  minus intercept (post fire weight)
    diffWeight = prefireWeight - postfireWeight
    #y = a * x ^power
    #t = (gram/a)^-power
    dateMinutes = pow((diffWeight / alpha), (1 / power))
    #dateMinutes = pow((diffWeight/alpha),power)
    dateYears, dateCalendar, ADBC = minToYears(dateMinutes)
    total_error = 0
    return dateYears, dateCalendar, ADBC, total_error


def age_output_nlin( power, power_stderr, alpha, alpha_stderr, prefireWeightAverage, prefireWeightStdDev,
                     postfireWeightAverage, postfireWeightStdDev):
    dateYears, dateCalendar, ADBC, dateErrorYears = age_calculate_nlin(power, power_stderr, alpha, alpha_stderr,
                                                                       postfireWeightAverage,
                                                                       postfireWeightStdDev, prefireWeightAverage,
                                                                       prefireWeightStdDev)
    results = ""
    results += "Difference in Absolute Weight (105 - 500):  " + str(prefireWeightAverage - postfireWeightAverage) + "\n"
    results += "alpha value:                                " + str(alpha) + "\n"
    results += "power value:                                " + str(power) + "\n"
    t = (dateYears, dateErrorYears)
    results += "Absolute Weight Change Age (Years):         %d +/- %0.4f \n" % t
    t = (ADBC, dateCalendar, dateErrorYears)
    results += "Absolute Weight Change Date:                %s %d +/- %0.4f \n" % t
    results += "------------------------------------------------------------------------------------------" + "\n"
    return results


def age_output( power, slope, intercept, r_value, p_value, std_err, ffilename, prefireWeightAverage,
                prefireWeightStdDev, postfireWeightAverage, postfireWeightStdDev, weightChange, percentWeightChange,
                runDate, sampleName):
    dateYears, dateCalendar, ADBC, dateErrorYears, slopeFourthPower = age_calculate(power, slope, std_err,
                                                                                    intercept,
                                                                                    postfireWeightAverage,
                                                                                    postfireWeightStdDev,
                                                                                    prefireWeightAverage,
                                                                                    prefireWeightStdDev)

    ## for just excel files -- save the results
    ## probably should do this for the sqlite too (but later..)
    if ExcelOrRHX == 1:
        ws.cell('H10').value = "Absolute Weight Slope (g/m^0.25)"
        ws.cell('I10').value = slope
        ws.cell('H11').value = "Absolute Weight Slope (g/m)"
        ws.cell('I11').value = slopeFourthPower
        ws.cell('H12').value = "Absolute WeightSlope StdErr"
        ws.cell('I12').value = std_err
        ws.cell('H13').value = "Absolute Weight Slope R-value"
        ws.cell('I13').value = r_value
        ws.cell('H14').value = "Age (BP)"
        ws.cell('I15').value = dateYears
        ws.cell('H16').value = "Age (AD/BC)"
        ws.cell('I17').value = dateCalendar
        ws.cell('J18').value = ADBC
        ws.cell('H19').value = "Error"
        ws.cell('I19').value = dateErrorYears
        wb.save(filename=ffilename)

    results = ""
    results += "Difference in Absolute Weight (105 - 500):  " + str(prefireWeightAverage - postfireWeightAverage) + "\n"
    results += "Power used in linear regression:            1/%0.2f \n" % power
    results += "RHX Slope (g/m^1/4):                        " + str(slope) + "\n"
    results += "RHX Slope (g/m):                            " + str(slopeFourthPower) + "\n"
    results += "RHX Intercept (g):                          " + str(intercept) + "\n"
    results += "RHX Slope r-value:                          " + str(r_value) + "\n"
    results += "RHX Slope std_err:                          " + str(std_err) + "\n"
    results += "Absolute Weight Change Age (Years):         " + str(dateYears) + " +/- " + str(dateErrorYears) + "\n"
    results += "Absolute Weight Change Date:                " + str(abs(dateCalendar)) + " " + str(ADBC) + " +/- "\
               + str(dateErrorYears) + "\n"
    results += "------------------------------------------------------------------------------------------" + "\n"

    return results


def markEvenlySpacedPoints(sx, sy):
    nextline = 0
    #sy = smooth(y)
    markcount = 0
    skipcount = 1
    evenlySpacedXArray = np.array([])
    evenlySpacedYArray = np.array([])

    for index, value in enumerate(sx):
        if index == 0 or index == nextline:
            markcount += 1
            evenlySpacedXArray = np.append(evenlySpacedXArray, value)
            evenlySpacedYArray = np.append(evenlySpacedYArray, sy[index])
            nextline = int(index + int(pow(skipcount, 1.25)))
            skipcount += 1

    return evenlySpacedXArray, evenlySpacedYArray


def getTempAndHumidityForLocation():
    temp = 0.0
    humidity = 0.0
    humidity = 0.0
    try:
        c = conn.cursor()
        c.execute('select f_temperature, f_humidity from tblRun where i_runID=1')
        row = c.fetchone()
        temp = row[0]
        humidity = row[1]
        #print "max date:", maxDate
    except sqlite.OperationalError, msg:
        logger.error("A SQL error occurred: %s", msg)
    return temp, humidity


def getSampleInfo(sampleID):
    row = []
    try:
        c = conn.cursor()
        logger.debug(
            "select t_sampleName, t_assemblageName from tblSample where i_positionNumber = %d" % sampleID)
        c.execute(
            "select  t_sampleName, t_assemblageName from tblSample where i_positionNumber = %d" % sampleID)
        row = c.fetchone()
        return row
    except sqlite.OperationalError, msg:
        logger.error("A SQL error occurred: %s" % msg)
        print "error: ", msg
        sys.exit()


def getRunInfo(runID):
    locationCode = ""
    numberOfSamples = 0
    description = ""
    temperature = 0.0
    humidity = 0.0
    status = True
    row = []
    try:
        t = int(runID, )
        c = conn.cursor()
        logger.debug(
            "select v_operatorName, v_locationCode, i_numberOfSamples, t_assemblageName, f_locationTemperature, f_locationHumidity, d_endOfFiring,i_durationOfFiring, t_notes, v_status from tblRun where i_runID = %d" % (
                runID))
        c.execute(
            'select v_operatorName, v_locationCode, i_numberOfSamples, t_assemblageName, f_locationTemperature, f_locationHumidity, d_endOfFiring,i_durationOfFiring, t_notes, v_status from tblRun where i_runID=%s' % runID)
        row = c.fetchone()
        return row
    except sqlite.OperationalError, msg:
        logger.error("A SQL error occurred: %s" % msg)
        print "error: ", msg
        sys.exit()


def closeDatabase():
# Ensure variable is defined
    try:
        conn
    except NameError:
        conn = None
    if conn is not None:
        try:
            conn.commit()
            conn.close()
            return True
        except sqlite.OperationalError, msg:
            logger.error("A SQL error occurred: %s", msg)
            return False
    return True

#def nlinRegression(timeArray,weightChangeArray, minval, maxval, nlinSubPlot):
#   r = robjects.r
#   r=X11()
#   nlx=[]
#   nly=[]
#   scount=0
#   count=0

#   for var in timeArray:
#      if var>minval and var<maxval:
#         nlx.append(var)
#         nly.append(weightChangeArray[count])
#         scount+=1
#      count += 1

#   x=FloatVector(nlx)
#   y=FloatVector(nly)
#   LM_fit5 = minpack.nlsLM(y ~ I(l * x^n), start=list(l=0.000661, n=x))
#   r.plot(r.runif(10), y, xlab="time (min)", ylab="weight change")
#   r.plot(x,r.fitted(LM_fit5), col='red')
#   r.summary(LM_fit5)
def aggregateAverages():
    ffilename = easygui.fileopenbox(msg='SQLLite Filename', title='select file', filetypes=['*.sqlite'],
                                    default=directory)
    if ffilename is None:
        sys.exit()
    global conn
    conn = sqlite.connect(ffilename, detect_types=sqlite.PARSE_DECLTYPES | sqlite.PARSE_COLNAMES)
    try:
        c = conn.cursor()
        sql = "DROP TABLE if exists tblAggregateMeasurement"
        c.execute(sql)
    except sqlite.OperationalError, msg:
        logger.error("A SQL error occurred: %s", msg)

    try:
        c = conn.cursor()
        sql = 'CREATE TABLE "tblAggregateMeasurement" ("i_measurementID" INTEGER PRIMARY KEY  NOT NULL ,"i_sampleID" INTEGER,"i_positionNumber" INTEGER,"i_preOrPost" INTEGER,"f_weight" DOUBLE,"d_dateTime" DATETIME,"f_elapsedTimeMin" DOUBLE,"f_temperature" FLOAT,"i_runID" INTEGER,"f_standardWeight" DOUBLE,"f_elapsedTimeQuarterPower" DOUBLE,"i_count" INTEGER,"v_status" VARCHAR,"f_humidity" FLOAT,"i_repetition" INTEGER,"i_repetitionCount" INTEGER);'
        c.execute(sql)
    except sqlite.OperationalError, msg:
        logger.error("A SQL error occurred: %s", msg)

    operator, location, numberOfSamples, assemblage, temperatureLocation, humidityLocation, dateOfRun, i_durationOfFiring, notes, status = getRunInfo(
        1)
    sampleNumber = 0
    measurementNumber = 0
    maxReps = 0
    while sampleNumber <= numberOfSamples:
        sampleNumber += 1
        try:
            c.execute("select max(i_repetition) as reps from tblMeasurement where i_positionNumber=%d" % sampleNumber)
            row = c.fetchone()
            maxReps = row[0]
        except sqlite.OperationalError, msg:
            logger.error("A SQL error occurred: %s", msg)
        aveTime = 0
        aveTimeQP = 0
        rep = 0
        while rep <= maxReps:
            rep += 1
            t = (sampleNumber, rep)
            try:
                c.execute(
                    "select avg(f_weight) as aveWeight, avg(f_elapsedTimeMin) as aveTime, avg(f_elapsedTimeQuarterPower) as avgTimeQP, avg(f_standardWeight) as aveStandardWeight, avg(f_temperature) as aveTemp, avg(f_humidity) as aveHumidity from tblMeasurement where i_positionNumber = ? and i_preOrPost=4 and i_repetition=?",
                    t)
                row = c.fetchone()
                aveWeight = row[0]
                aveTime = row[1]
                aveTimeQP = row[2]
                aveStandardWeight = row[3]
                aveTemp = row[4]
                aveHumidity = row[5]
            except sqlite.OperationalError, msg:
                logger.error("A SQL error occurred: %s", msg)
            dateNow = datetime.today()
            measurementNumber += 1
            if aveWeight > 0:
                try:
                    t = (measurementNumber, sampleNumber, sampleNumber, 4, aveWeight, dateNow, aveTime, aveTimeQP,
                         aveStandardWeight, aveTemp, aveHumidity, rep)
                    c.execute(
                        "insert into tblAggregateMeasurement (i_measurementID, i_sampleID, i_positionNumber, i_preOrPost, f_weight, d_dateTime,f_elapsedTimeMin,f_elapsedTimeQuarterPower, f_standardWeight, f_temperature,f_humidity,i_repetition) VALUES (?,?,?,?,?,?,?,?,?,?,?,?) ",
                        t)
                    conn.commit()
                except sqlite.OperationalError, msg:
                    logger.error("A SQL error occurred: %s", msg)

    print "Done! Now do the analysis."
    closeDatabase()
    return ffilename


def exportData():
    ffilename = easygui.fileopenbox(msg='SQLLite Filename', title='select file', filetypes=['*.sqlite'],
                                    default=directory)
    if ffilename is None:
        sys.exit()
    newfilename = easygui.filesavebox(msg='Excel File', title='Create a new excel file to save the output',
                                      filetypes=['*.xlsx'], default=directory)
    if newfilename is None:
        sys.exit()
    wb = Workbook()
    global conn
    conn = sqlite.connect(ffilename, detect_types=sqlite.PARSE_DECLTYPES | sqlite.PARSE_COLNAMES)

    operator, location, numberOfSamples, assemblage, temperatureLocation, humidityLocation, dateOfRun, i_durationOfFiring, notes, status = getRunInfo(
        1)

    row = []
    print "Number of samples to export: ", numberOfSamples
    sampleNumber = 0
    worksheetNumber = -1

    while sampleNumber < numberOfSamples:
        worksheetNumber += 1
        sampleNumber += 1
        sampleName, assemblageName = getSampleInfo(sampleNumber)
        print "Exporting Sample:  ", sampleName
        if sampleNumber > 1:
            ws = wb.create_sheet()

        ws = wb.worksheets[worksheetNumber]
        ws.title = assemblage + " - " + sampleName
        ws.cell('A1').value = "Elapsed Time (min)"
        ws.cell('B1').value = "Elapsed Time (min^1/4)"
        ws.cell('C1').value = "Weight (g)"
        ws.cell('D1').value = "Temperature (C)"
        ws.cell('E1').value = "Humidity (%rh)"
        ws.cell('F1').value = "Standard Weight (g)"
        rn = 1
        try:
            c = conn.cursor()
            c.execute(
                'select  f_weight, f_elapsedTimeMin, f_elapsedTimeQuarterPower, f_temperature, f_humidity, f_standardWeight from tblMeasurement where '
                'i_positionNumber=%d and i_preOrPost=4  ' % sampleNumber)
            for row in c.fetchall():
                rn += 1
                weight = row[0]
                elapsedTimeMin = row[1]
                elapsedTimeQuarterPower = row[2]
                temperature = row[3]
                humidity = row[4]
                standardWeight = row[5]
                rowNumber = str(rn)
                ws.cell("A" + rowNumber).value = elapsedTimeMin
                ws.cell("B" + rowNumber).value = elapsedTimeQuarterPower
                ws.cell("C" + rowNumber).value = weight
                ws.cell("D" + rowNumber).value = temperature
                ws.cell("E" + rowNumber).value = humidity
                ws.cell("F" + rowNumber).value = standardWeight
        except sqlite.OperationalError, msg:
            logger.error("A SQL error occurred in export Data: %s", msg)
            sys.exit()

    wb.save(filename=newfilename)
    closeDatabase()
    easygui.msgbox(msg="Export complete!")


def getMeanTempAndHumidtyForSample(sampleNumber):
    try:
        c = conn.cursor()
        c.execute('select  avg(f_temperature), avg(f_humidity) from tblMeasurement where '
                  'i_positionNumber=%d and i_preOrPost=4  ' % sampleNumber)
        row = c.fetchone()
        aveTemperature = row[0]
        aveHumidity = row[1]
        return aveTemperature, aveHumidity
    except sqlite.OperationalError, msg:
        logger.error("A SQL error occurred in getMeanTemperatureOfRunForSample: %s", msg)
        sys.exit()


def askForELT(sampleNumber, meanTemp):
    msg = "The mean temperature for sample: " + str(sampleNumber) + " has been " + str(
        meanTemp) + ". Enter ELT for sample: " + str(sampleNumber)
    ELT = 0  # we start with blanks for the values
    ELT = easygui.enterbox(msg, "Temperature Entry")
    return ELT

############################################################################################
############################################################################################


logger = logging.getLogger("rhxAnalyze-Combined")
logger.setLevel(logging.DEBUG)
# create file handler which logs even debug messages
fh = logging.FileHandler('analyzeCombined.log')
fh.setLevel(logging.DEBUG)
# create console handler with a higher log level
ch = logging.StreamHandler()
ch.setLevel(logging.ERROR)
# create formatter and add it to the handlers
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
fh.setFormatter(formatter)
ch.setFormatter(formatter)
# add the handlers to the logger
logger.addHandler(fh)
logger.addHandler(ch)

directory = ''
if os.name == 'nt':
    directory = 'c:\\Users\\Archy\\Dropbox\\Rehydroxylation\\'
elif os.name == 'mac':
    directory = '/Users/clipo/Dropbox/Rehydroxylation/'

numberOfSamples = 0
postfireWeightAverage = 0.0
postfireWeightStdDev = 0.0
prefireWeightAverage = 0.0
prefireWeightStdDev = 0.0
percentWeightChange = 0.0
weightChange = 0.0
operator = str
location = str
numberOfSamples = 0
assemblage = str
temperatureLocation = str
humidityLocation = str
dateOfRun = datetime
i_durationOfFiring = 0
notes = str
status = str
delay = 0
runDate = datetime
ffilename = str

while 1:
    if DEBUG == 1:
        ExcelOrRHX = 1
    else:
        ExcelOrRHX = ask_for_excel_or_sqlite()

    if ExcelOrRHX == 0:
        sys.exit()
    elif ExcelOrRHX == 5:
        updateAndFixAverageValues()
    elif ExcelOrRHX == 4:
        # create aggregate averages
        ffilename = aggregateAverages()
        numberOfSamples = 0
        try:
            c = conn.cursor()
            logger.debug('select i_numberOfSamples from tblRun')
            c.execute('select i_numberOfSamples from tblRun')
            for row in c.fetchall():
                ## numberOfSamples=row["i_numberOfSamples"]
                numberOfSamples = row[0]
        except sqlite.OperationalError, msg:
            logger.error("A SQL error occurred: %s", msg)
            sys.exit()
        operator, location, numberOfSamples, assemblage, temperatureLocation, humidityLocation, dateOfRun, i_durationOfFiring, notes, status = getRunInfo(
            1)
        if status != "postfire":
            easygui.msgbox(msg="The dataset must have post-fire RHX measurements to continue.")
            closeDatabase()
            sys.exit()
        runDate = datetime.strptime(str(dateOfRun), '%m-%d-%y %H:%M:%S')
        delay = datetime.today() - runDate
    elif ExcelOrRHX == 3:
        ## export file
        exportData()
        sys.exit()
    elif ExcelOrRHX == 1:
        if DEBUG == 1:
            ffilename = "/Users/clipo/Dropbox/Rehydroxylation/VTI-SA Data/LB0055-Run1.xlsx"
        else:
            ffilename = easygui.fileopenbox(msg='Excel File', title='select file', filetypes=['*.xlsx'],
                                            default=directory)
        if ffilename is None:
            sys.exit()
        wb = load_workbook(filename=ffilename)
        listOfWorksheets = wb.get_sheet_names()  ## get the list of the sheets in the workbook
        worksheet = ask_for_worksheet(listOfWorksheets)  ## ask which sheet we want to use
        ws = wb.get_sheet_by_name(name=worksheet)  ## go get the sheet
        sampleName = ws.cell('C5').value  ## getting values in standardized places
        operator = ws.cell('C3').value
        dateOfRun = ws.cell('C15').value
        try:
            runDate = datetime.strptime(str(dateOfRun), '%m-%d-%Y')
        except ValueError:
            runDate = datetime.today()
        notes = str(ws.cell('C7').value) + " " + str(ws.cell('C8').value)

        ## first check for weight values
        originalWeighAverage = ws.cell('I1').value

        ## if no weight values --- ask for them and enter them into spreadsheet
        if originalWeighAverage is None:
            originalWeightAverage, originalWeightStdDev, prefireWeightAverage, prefireWeightStdDev, postfireWeightAverage, postfireWeightStdDev = ask_for_weight_values()

            ## enter the values into the Excel Spreadsheet
            ws.cell('H1').value = "Original Weight (g)"
            ws.cell('I1').value = originalWeightAverage
            ws.cell('H2').value = "Original Weight StdDev)"
            ws.cell('I2').value = originalWeightStdDev
            ws.cell('H3').value = "105 Degree C Weight (g)"
            ws.cell('I3').value = prefireWeightAverage
            ws.cell('H4').value = "105 Degree C Weight StdDev"
            ws.cell('I4').value = prefireWeightStdDev
            ws.cell('H5').value = "500 Degree C Weight (g)"
            ws.cell('I5').value = postfireWeightAverage
            ws.cell('H6').value = "500 Degree C Weight StdDev"
            ws.cell('I6').value = postfireWeightStdDev

            weightChange = prefireWeightAverage - postfireWeightAverage
            percentWeightChange = weightChange / prefireWeightAverage
            ws.cell('H7').value = "Weight Change (g)"
            ws.cell('I7').value = weightChange
            ws.cell('H8').value = "Percent Weight Change (%)"
            ws.cell('I8').value = percentWeightChange

            ## delay since firing
            delay = float(ask_for_time_delay())
            if delay == "":
                delay = 0.0
            ws.cell('H9').value = "Delay Since Firing (minutes)"
            ws.cell('I9').value = delay
            ws.cell('H10').value = "Date of Analysis"
            now = datetime.today()
            today = now.strftime("%m-%d-%y %H:%M:%S")
            ws.cell('I10').value = today

        ## in this case read from the spreadsheet.
        else:
            fileName = ws.cell('A1').value
            originalWeightAverage = float(ws.cell('I1').value)
            originalWeightStdDev = float(ws.cell('I2').value)
            prefireWeightAverage = float(ws.cell('I3').value)
            prefireWeightStdDev = float(ws.cell('I4').value)
            postfireWeightAverage = float(ws.cell('I5').value)
            postfireWeightStdDev = float(ws.cell('I6').value)
            weightChange = float(ws.cell('I7').value)
            percentWeightChange = float(ws.cell('I8').value)
            delay = float(ws.cell('I9').value)

        #print "delay:  ", delay
        ##THERE CAN ONLY BE ONE SAMPLE FOR THE EXCEL CASE -- NOT SO FOR THE RHX CHAMBER
        numberOfSamples = 1
        wb.save(filename=ffilename)
        ## alternative is that its a sqlite file -- RHX chamber
    elif ExcelOrRHX == 2:
        ffilename = easygui.fileopenbox(msg='SQLLite Filename', title='select file', filetypes=['*.sqlite'],
                                        default=directory)
        if ffilename is None:
            sys.exit()
        conn = sqlite.connect(ffilename, detect_types=sqlite.PARSE_DECLTYPES | sqlite.PARSE_COLNAMES)
        numberOfSamples = 0
        try:
            c = conn.cursor()
            logger.debug('select i_numberOfSamples from tblRun')
            c.execute('select i_numberOfSamples from tblRun')
            for row in c.fetchall():
                ## numberOfSamples=row["i_numberOfSamples"]
                numberOfSamples = row[0]
        except sqlite.OperationalError, msg:
            logger.error("A SQL error occurred: %s", msg)
            sys.exit()
        operator, location, numberOfSamples, assemblage, temperatureLocation, humidityLocation, dateOfRun, i_durationOfFiring, notes, status = getRunInfo(
            1)
        if status != "postfire":
            easygui.msgbox(msg="The dataset must have post-fire RHX measurements to continue.")
            closeDatabase()
            sys.exit()
        runDate = datetime.strptime(str(dateOfRun), '%m-%d-%y %H:%M:%S')
        delay = datetime.today() - runDate
        print "Number of samples: ", numberOfSamples

    ## for each sample get the data for the post-fire
    sampleNumber = 1

    while sampleNumber <= numberOfSamples:
        sampleName = "1"
        assemblage = ""
        if ExcelOrRHX == 2:
            sampleName, assemblageName = getSampleInfo(sampleNumber)
            print "Location: ", location
            print "Assemblage: ", assemblageName
            print "Sample Number: ", sampleName
            print "Operator:  ", operator
            print "Run Date: ", dateOfRun
            print "Notes:  ", notes
            print "---------------------------------------"

        #humidity, coefficient = ask_for_RHCoefficient()
        params = {'figure.figsize': [8, 8], 'text.fontsize': 8, 'xtick.labelsize': 8, 'ytick.labelsize': 8,
                  'axes.fontsize': 8, 'axes.style': 'plain', 'axes.axis': 'both'}
        plt.rcParams.update(params)
        fig = plt.figure(sampleNumber)
        fig.suptitle(sampleName, fontsize=14)
        fig.set_size_inches(18, 18)
        ### HERE WE CREATE THE PLOT OBJECTS
        ## define all of the plots in the figure
        fontsize = 8
        ax1 = fig.add_subplot(6, 3, 1)                    # weight change over time^1/4
        ax1.set_title('Weight Change/Time^0.25', fontsize=fontsize)
        weightRegression = fig.add_subplot(6, 3, 2)          ## weight change that will be regression
        weightRegression.set_title("Linear regression for weight change", fontsize=fontsize)
        weightRegression.ticklabel_format(useOffset=False)
        tempPlot = fig.add_subplot(6, 3, 3)               # filtered weight graph
        tempPlot.set_title("Temperature (C)", fontsize=fontsize)
        tempPlot.ticklabel_format(useOffset=False)
        humidPlot = fig.add_subplot(6, 3, 6)              # the humidity graph
        humidPlot.set_title("Humidity (%RH)", fontsize=fontsize)
        humidPlot.ticklabel_format(useOffset=False)
        slopeGraph = fig.add_subplot(6, 3, 9)
        slopeGraph.set_title("Moving window slope value over time", fontsize=fontsize)
        slopeGraph.ticklabel_format(useOffset=False)
        filteredRHPlot = fig.add_subplot(6, 3, 10)
        filteredRHPlot.set_title("Weight change filtered by mean RH% +/- 1", fontsize=fontsize)
        filteredRHPlot.ticklabel_format(useOffset=False)
        filteredRHRegressionPlot = fig.add_subplot(6, 3, 11)
        filteredRHRegressionPlot.set_title("Weight change regression filtered by RH%", fontsize=fontsize)
        filteredRHRegressionPlot.ticklabel_format(useOffset=False)
        dateGraph = fig.add_subplot(6, 3, 12)
        dateGraph.set_title("Moving window date calculation over time", fontsize=fontsize)
        dateGraph.ticklabel_format(useOffset=False)
        evenPlot = fig.add_subplot(6, 3, 4)
        evenPlot.set_title("Weight change over evenly spaced points in time (^0.25)", fontsize=fontsize)
        evenPlot.ticklabel_format(useOffset=False)
        evenPlotRegression = fig.add_subplot(6, 3, 5)
        evenPlotRegression.set_title("Linear regression for weight change over evenly spaced points.",
                                     fontsize=(fontsize - 1))
        evenPlotRegression.ticklabel_format(useOffset=False)
        weightPlotWithLinearTime = fig.add_subplot(6, 3, 7)
        weightPlotWithLinearTime.set_title("Weight change over linear time", fontsize=fontsize)
        weightPlotWithLinearTime.ticklabel_format(useOffset=False)
        nlinRegressionPlot = fig.add_subplot(6, 3, 8)
        nlinRegressionPlot.set_title("Nonlinear regression for weight change over linear time", fontsize=fontsize)
        nlinRegressionPlot.ticklabel_format(useOffset=False)
        nlinRegressionFitPlot = fig.add_subplot(6, 3, 13)
        nlinRegressionFitPlot.set_title("Nonlinear regression fit ", fontsize=fontsize)
        nlinRegressionFitPlot.ticklabel_format(useOffset=False)

        multiLinRegressionFitPlot = fig.add_subplot(6, 3, 14)
        multiLinRegressionFitPlot.set_title("Multiple linear regression fit ", fontsize=fontsize)
        multiLinRegressionFitPlot.ticklabel_format(useOffset=False)

        multiLinRegressionFitPlotB = fig.add_subplot(6, 3, 15)
        multiLinRegressionFitPlotB.set_title("Multiple linear regression fit ", fontsize=fontsize)
        multiLinRegressionFitPlotB.ticklabel_format(useOffset=False)

        #fig.tight_layout()
        fig.subplots_adjust(hspace=.5, wspace=.5)
        ## only plot the standard weight if we are dealing with the RHX chamber
        if ExcelOrRHX == 2:
            standardWeightPlot = fig.add_subplot(6, 3, 16)
            standardWeightPlot.set_title("Weight of standard over time", fontsize=fontsize)
            standardWeightPlot.ticklabel_format(useOffset=False)

        ## dont plot percentages unless we ask (see constants)
        if CALCULATE_AGE_FROM_PERCENTS is True:
            weightPercentPlotWithLinearTime = fig.add_subplot(6, 3, 18)
            weightPercentPlotWithLinearTime.ticklabel_format(useOffset=False)
            #nlinSubPlot = fig.add_subplot(4,3,10)

        finish = 0
        preFireWeight = 0.0
        postFireWeight = 0.0
        x = np.array([])
        y = np.array([])
        ypercent = np.array([])
        temp = np.array([])
        humid = np.array([])
        tempSmooth = np.array([])
        humidSmooth = np.array([])
        y_correct = np.array([])
        row_list = np.array([])
        standardWeight = np.array([])
        standardWeightNonZero = np.array([])
        standardWeightNonZeroTime = np.array([])
        timeArray = np.array([])
        weightChangeArray = np.array([])
        humidAverage = 0.0
        humidStdDev = 0.0
        tempAverage = 0.0
        tempStdDev = 0.0
        rowNumber = 20
        endOfData = 0
        initialWeight = 0.0
        tempSum = 0.0
        weightChangeValue = 0.0
        ## in the Excel Case
        ## read in the data
        if ExcelOrRHX == 1:
            count = 0
            rowNumber = 20
            initialWeight = float(ws.cell('C' + str(rowNumber)).value)
            meanTemp = 0.0
            while endOfData == 0:
                count += 1
                xv = ws.cell('A' + str(rowNumber)).value
                if xv is None:
                    break
                xv = float(xv) + delay
                #print xv
                x = np.append(x, (pow(float(xv), 0.25)))
                weightChangeValue = float(ws.cell('B' + str(rowNumber)).value / 1000) - initialWeight
                y = np.append(y, weightChangeValue) # turn to grams
                ypercent = np.append(ypercent, (float(ws.cell('C' + str(rowNumber)).value)))
                temp = np.append(temp, ws.cell('D' + str(rowNumber)).value)
                tempSum += ws.cell('D' + str(rowNumber)).value
                humid = np.append(humid, (ws.cell('E' + str(rowNumber)).value))
                timeArray = np.append(timeArray, xv)
                weightChangeArray = np.append(weightChangeArray, weightChangeValue)
                meanTemp = float(tempSum / count)
                #print "mean temperature: ", meanTemp, " tempSum: ", tempSum, " count: ", count
                rowNumber += 1
                #print "X :", x[count-1], " Y: ", y[count -1]
        else:
            ### the SQLITE case
            rowcount = 0
            try:
                c = conn.cursor()
                c.execute('select f_preHumidityAverage, f_preHumidityStdDev,f_preTempAverage, '
                          'f_preTempStdDev,f_preWeightAverage,f_preWeightStdDev,f_postFireWeightAverage,f_postFireWeightStdDev '
                          'from tblSample where i_positionNumber = %d' % sampleNumber)
                for row in c.fetchall():
                    humidAverage = row[0]
                    humidStdDev = row[1]
                    tempAverage = row[2]
                    tempStdDev = row[3]
                    prefireWeightAverage = row[4]
                    prefireWeightStdDev = row[5]
                    postfireWeightAverage = row[6]
                    postfireWeightStdDev = row[7]
                    weightChange = prefireWeightAverage - postfireWeightAverage
                    percentWeightChange = (prefireWeightAverage - postfireWeightAverage) / prefireWeightAverage

            except sqlite.OperationalError, msg:
                logger.error("A SQL error occurred: %s", msg)
                sys.exit()
            meanTemp, meanHumidity = getMeanTempAndHumidtyForSample(sampleNumber)
            time = []
            humidMax = 0
            humidMin = 0
            tempMax = 0
            tempMin = 0
            count = 0
            if humidAverage is None:
                humidMax = 0
                humidMin = 0
            else:
                humidMax = humidAverage + humidStdDev
                humidMin = humidAverage - humidStdDev
            if tempAverage is None:
                tempMax = 0
                tempMin = 0
            else:
                tempMax = tempAverage + tempStdDev
                tempMin = tempAverage - tempStdDev

            try:
                c.execute('select min(f_weight) as MIN, max(f_weight) as MAX from tblMeasurement where '
                          'i_positionNumber=%d and '
                          'i_preOrPost=4 ' % sampleNumber)
                row = c.fetchone()
                weightMin = row[0]
                weightMax = row[1]
                #print "minimum weight: ", weightMin
            except sqlite.OperationalError, msg:
                logger.error("A SQL error has occurred in selecting min and max weights: %s", msg)
            try:
                c = conn.cursor()
                if ExcelOrRHX == 4:
                    c.execute(
                        'select  f_weight, f_elapsedTimeQuarterPower, f_temperature, f_humidity, f_standardWeight, f_elapsedTimeMin from tblAggregateMeasurement where '
                        'i_positionNumber=%d and i_preOrPost=4  ' % sampleNumber)
                    logger.debug(
                        'select  f_weight, f_elapsedTimeQuarterPower, f_temperature, f_humidity, f_standardWeight, f_elapsedTimeMin from tblAggregateMeasurement where '
                        'i_positionNumber=%d and i_preOrPost=4  ' % sampleNumber)
                else:
                    c.execute('select  f_weight, f_elapsedTimeQuarterPower, f_temperature, f_humidity, '
                              'f_standardWeight, f_elapsedTimeMin from tblMeasurement where '
                              'i_positionNumber=%d and i_preOrPost=4  ' % sampleNumber)
                    logger.debug('select  f_weight, f_elapsedTimeQuarterPower, f_temperature, f_humidity, '
                                 'f_standardWeight, f_elapsedTimeMin from tblMeasurement where '
                                 'i_positionNumber=%d and i_preOrPost=4  ' % sampleNumber)

                for row in c.fetchall():
                    count += 1
                    if count == 1:
                        initialWeight = float(row[0])
                    if count > 0 and float(row[0]) > 0.250:
                        row_list = np.append(row_list, count)
                        x = np.append(x, float(row[1]))
                        #print "x: ", x
                        weightChangeArray = np.append(weightChangeArray, (float(row[0]) - weightMin))
                        y = np.append(y, (float(row[0]) - weightMin))
                        temp = np.append(temp, float(row[2]))
                        humid = np.append(humid, float(row[3]))
                        time = np.append(time, float(row[5]))
                        standardWeight = np.append(standardWeight, float(row[4]))
                        if float(row[4]) > 0:
                            standardWeightNonZero = np.append(standardWeightNonZero, float(row[4]))
                            standardWeightNonZeroTime = np.append(standardWeightNonZeroTime, float(row[1]))
                        ypercent = np.append(ypercent, ((float(row[0]) - weightMin) / weightMin))
                        timeArray = np.append(timeArray, float(row[5]))
            except sqlite.OperationalError, msg:
                logger.error("A SQL error has occurred: %s", msg)
                ################## END OF SQLITE STUFF ####################


        #smooth_weightChangeArray = ds.smooth_data (x, weightChangeArray, d=4, stdev=1e-1)
        smooth_y = smoothTriangle(y, 10)

        if ExcelOrRHX == 2:
            standardWeightPlot.plot(standardWeightNonZeroTime, standardWeightNonZero, '.')
            standardWeightPlot.set_xlabel('Time (m^1/4)', fontsize=8)
            standardWeightPlot.set_ylabel('Weight of Standard (g)', fontsize=8)
            standardWeightPlot.ticklabel_format(style='plain', axis='both', fontsize=8)
            standardWeightPlot.grid(True)

        weightPlotWithLinearTime.plot(timeArray, smooth_y, '.')
        weightPlotWithLinearTime.set_xlabel('Time (m)', fontsize=8)
        weightPlotWithLinearTime.set_ylabel('Weight (g)', fontsize=8)
        weightPlotWithLinearTime.ticklabel_format(style='plain', axis='both', fontsize=8)
        weightPlotWithLinearTime.grid(True)

        if CALCULATE_AGE_FROM_PERCENTS is True:
            weightPercentPlotWithLinearTime.plot(timeArray, ypercent, '.')
            weightPercentPlotWithLinearTime.set_xlabel('Time (m)', fontsize=8)
            weightPercentPlotWithLinearTime.set_ylabel('Weight Percent (g)', fontsize=8)
            weightPercentPlotWithLinearTime.ticklabel_format(style='plain', axis='both', fontsize=8)
            weightPercentPlotWithLinearTime.grid(True)

        #x, y = smooth(x, y)
        ## now create an array that has the evenly spaced values (and update the sql table with these values)
        evenlySpacedYPercentArray = np.array([])
        evenlySpacedXArray = np.array([])
        evenlySpacedYArray = np.array([])
        evenlySpacedXArray, evenlySpacedYArray = markEvenlySpacedPoints(x, smooth_y)
        evenlySpacedXPercentArray, evenlySpacedYPercentArray = markEvenlySpacedPoints(x, ypercent)
        evenlySpacedPlot(evenPlot, evenlySpacedXArray, evenlySpacedYPercentArray)

        overall_plot(ax1, x, smooth_y, sampleNumber)
        temp_humidity_plot(tempPlot, humidPlot, x, temp, humid)

        ## iterative
        step = 100
        count = 0
        steps = int(np.size(x) / step)
        #print "Number of steps to take: ", steps
        residualYArray = np.array([])
        residualXArray = np.array([])
        powerYArray=np.array([])
        alphaYArray=np.array([])
        while count < steps * step:
            # create a set of Parameters
            params = Parameters()
            params.add('alpha', value=1)
            #params.add('power', value=0.25)

            # do fit, here with leastsq model
            result = minimize(quarterPowerFunction, params, args=(x[count:], y[count:]))
            residualXArray = np.append(residualXArray, x[count])
            stdDevOverMean = np.std(result.residual) / np.average(result.residual)
            residualYArray = np.append(residualYArray, stdDevOverMean)
            #powerYArray = np.append( powerYArray,  params.get('power').value)
            #print "Power: ", params.get('power')
            alphaYArray = np.append( alphaYArray, params.get('alpha').value)
            #print "Alpha: ", params.get('alpha')
            #print "Nlin regression step number: ", count
            count += step
        nlinRegressionFitPlot2 = nlinRegressionFitPlot.twinx()
        #nlinRegressionFitPlot.plot(residualXArray, powerYArray, 'o')
        nlinRegressionFitPlot2.plot(residualXArray, alphaYArray, 'x' )
        nlinRegressionFitPlot.plot(residualXArray, residualYArray, 'o')
        xls = (np.amin(residualXArray), np.amax(residualXArray))
        yls = (np.amin(residualYArray), np.amax(residualYArray))
        yls2 = (np.amin(alphaYArray), np.amax(alphaYArray))
        x1 = 0.0
        x2 = 0.0
        nlinRegressionFitPlot.axis([xls[0], xls[1], yls[0], yls[1]])
        nlinRegressionFitPlot.set_xlabel('Time (m)', fontsize=8)
        nlinRegressionFitPlot.set_ylabel('Std/Mean', fontsize=8)
        nlinRegressionFitPlot2.set_ylabel('alpha', fontsize=8)
        nlinRegressionFitPlot2.axis([xls[0], xls[1], yls2[0], yls2[1]])
        nlinRegressionFitPlot.ticklabel_format(style='plain', axis='both, fontsize=8')
        nlinRegressionFitPlot.grid(True)

        fig2 = plt.figure(sampleNumber)

        slopeArray = zeros(alen(x))
        interceptArray = zeros(alen(x))
        r_valueArray = zeros(alen(x))
        p_valueArray = zeros(alen(x))
        std_errArray = zeros(alen(x))
        dateArray = zeros(alen(x))
        percentSlopeArray = zeros(alen(x))

        filteredXarray = np.array([])
        filteredYarray = np.array([])
        filteredXarray, filteredYarray = filter_by_humidity(x, smooth_y, humid)

        filteredRHPlot.set_xlabel('Time (m^1/4', fontsize=8)
        filteredRHPlot.set_ylabel('Weight change (g)', fontsize=8)
        filteredRHPlot.ticklabel_format(style='plain', axis='both', fontsize=8)
        filteredRHPlot.grid(True)
        filteredRHPlot.plot(filteredXarray, filteredYarray, '.')

        ### size of the window to calculate slope, intercept, etc. ####
        window = 10
        #########################################

        (slopeArray, interceptArray, r_valueArray, p_valueArray, std_errArray, dateArray) = graph_slope(
            sampleNumber, evenlySpacedXArray, evenlySpacedYArray, window,
            prefireWeightAverage, prefireWeightStdDev, postfireWeightAverage, postfireWeightStdDev)

        dYSlope = np.array([])
        dYSlope = (np.roll(evenlySpacedYArray, -1) - evenlySpacedYArray)[:-1]
        dXSlope = np.array([])
        dXSlope = (np.roll(evenlySpacedXArray, -1) - evenlySpacedXArray)[:-1]
        slopeList = np.array([])
        slopeList = dYSlope/dXSlope

        slopeGraph.plot(evenlySpacedXArray[:-1], slopeList, '.')
        slopeGraph.set_xlabel('Time (m^1/4)', fontsize=8)
        slopeGraph.set_ylabel('Slope', fontsize=8)
        slopeGraph.ticklabel_format(style='plain', axis='both', fontsize=8)
        slopeGraph.grid(True)

        dateGraph.set_xlabel('Time (m^1/4)', fontsize=8)
        dateGraph.set_ylabel('Date (Years)', fontsize=8)
        dateGraph.ticklabel_format(style='plain', axis='both', fontsize=8)
        dateGraph.grid(True)
        dateGraph.plot(evenlySpacedXArray, dateArray, '--')

        #plt.plot(xArray,ainterceptArray,'x')
        plt.ion()     # turns on interactive mode
        plt.show()    # now this should be non-blocking\

        ## multiple linear regressions with different powers (0.3, 0.25, 0.2, 0.1, etc)

        powerTest = 0.3
        powerStep = -0.005
        powerArray = np.array([])
        weightArray = np.array([])
        slopeArray=np.array([])
        r_valueArray=np.array([])
        p_valueArray=np.array([])
        std_errArray=np.array([])
        powerValueArray=np.array([])
        while powerTest > 0:

            powerArray = pow(timeArray[700:],powerTest)
            weightArray = y[700:]
            slope, intercept, r_value, p_value, std_err = stats.linregress(powerArray, weightArray)
            powerValueArray=np.append(powerValueArray,1/powerTest)
            slopeArray=np.append(slopeArray, slope)
            r_valueArray=np.append(r_valueArray,r_value)
            p_valueArray=np.append(p_valueArray,p_value)
            std_errArray=np.append(std_errArray,std_err)
            powerTest += powerStep
        multiLinRegressionFitPlot2 = multiLinRegressionFitPlot.twinx()
        multiLinRegressionFitPlot.plot(powerValueArray, r_valueArray, 'x')

        dY = (np.roll(r_valueArray, -1) - r_valueArray)[:-1]
        dX = (np.roll(powerValueArray, -1) - powerValueArray)[:-1]
        slopes = dY/dX

        multiLinRegressionFitPlot2.plot(powerValueArray[:-1], slopes, '+', color='r' )
        multiLinRegressionFitPlot.set_xlabel('Power', fontsize=8)
        multiLinRegressionFitPlot.set_ylabel('r-value (x)', fontsize=8)
        multiLinRegressionFitPlot2.set_ylabel('Slopes of r-values (x)', fontsize=8)
        multiLinRegressionFitPlot.ticklabel_format(style='plain', axis='both, fontsize=8')
        multiLinRegressionFitPlot2.ticklabel_format(style='plain', axis='both, fontsize=8')
        #multiLinRegressionFitPlot.set_xscale('log')
        multiLinRegressionFitPlotB2 = multiLinRegressionFitPlotB.twinx()
        multiLinRegressionFitPlotB.plot(powerValueArray, slopeArray, 'x')
        multiLinRegressionFitPlotB2.plot(powerValueArray, std_errArray, '+', color='r' )
        multiLinRegressionFitPlotB.set_xlabel('Power', fontsize=8)
        multiLinRegressionFitPlotB.set_ylabel('Slope (x)', fontsize=8)
        multiLinRegressionFitPlotB2.set_ylabel('Std Error (+)', fontsize=8)
        multiLinRegressionFitPlotB.ticklabel_format(style='plain', axis='both, fontsize=8')
        multiLinRegressionFitPlotB2.ticklabel_format(style='plain', axis='both, fontsize=8')
        #multiLinRegressionFitPlotB.set_xscale('log')

        minval = ""
        maxval = ""
        results = ""
        asking = 0
        errormessage = ""
        minIndex = 0
        maxIndex = 0
        c = 0
        power=4
        while asking == 0:
            minval, maxval, power = ask_for_minmax_values(min(x), max(x), errormessage)
            sx = np.array([])
            sy = np.array([])
            sTimeArray = np.array([])
            ## chop out the data from the basic array (smoothed)
            for index, var in enumerate(x):
                if var >= minval and var <= maxval:
                    if c == 0:
                        minIndex = index
                        minX = var
                        minY = y[index]
                        minTimeArray = timeArray[index]
                    sx = np.append(sx, var)
                    sy = np.append(sy, y[index])
                    sTimeArray = np.append(sTimeArray, timeArray[index])
                    maxIndex = index
                    c += 1
            if np.alen(sx) < 4:
                errormessage = "Too few data points for evenly spaced regression. Must be at least 3. Try again. "
            else:
                asking = 1

        #chop out the data from the filtered array
        fx = np.array([])
        fy = np.array([])
        c=0
        for index, var in enumerate(filteredXarray):
            if var >= minval and var <= maxval:
                if c==0:
                    minX= var
                    minY=filteredXarray[index]
                fx = np.append(fx, var)
                fy = np.append(fy, filteredYarray[index])

        # chop out the data from the evenly spaced array
        ex = np.array([])
        ey = np.array([])
        ppx = np.array([])
        ppy = np.array([])
        c=0
        for index, var in enumerate(evenlySpacedXArray):
            if var >= minval and var <= maxval:
                if c==0:
                    minX=var
                    minY=evenlySpacedYArray[index]
                ex = np.append(ex, var)
                ey = np.append(ey, evenlySpacedYArray[index])
                ppx = np.append(ppx, var)
                ppy = np.append(ppy, evenlySpacedYPercentArray[index])

        ## from http://newville.github.com/lmfit-py/parameters.html
        # create a set of Parameters
        params = Parameters()
        params.add('alpha', value=1)
        params.add('power', value=0.25)
        # do fit, here with leastsq model
        result = minimize(fcn2min, params, method='leastsq', args=(sTimeArray, sy), scale_covar=True)
        #matrix_y = np.asmatrix(sy)
        # calculate final result
        final = sy + result.residual

        # write error report
        report_errors(params)
        p = params.get('power')
        nlinpower = float(p.value)
        #nlinpower = 0.25
        #power=4
        #power_stderr = float(p.stderr)
        power_stderr = 0.0
        a = params.get('alpha')
        alpha = float(a.value)
        alpha_stderr = float(a.stderr)
        # try to plot results
        nlinRegressionPlot.plot(sTimeArray, sy, 'k+')
        nlinRegressionPlot.plot(sTimeArray, final, 'r')

        results = ""
        results += "------------------------------------------------------------------------------------------\n"
        results += "Results for sample:                         " + str(sampleName) + " from file: " + ffilename + "\n"
        t = (minval, maxval)
        results += "Calculation based on date from: %0.3f (min^1/4)  To:  %0.3f (min^1/4)\n" % t
        results += "------------------------------------------------------------------------------------------\n"
        results += "Pre-fire (105) Absolute Weight (g):          " + str(prefireWeightAverage) + " +/- " + str(
            prefireWeightStdDev) + "\n"
        results += "Post-fire (500) Absolute Weight (g):         " + str(postfireWeightAverage) + " +/- " + str(
            postfireWeightStdDev) + "\n"
        results += "------------------------------------------------------------------------------------------\n"

        results += " ------ RESULTS BASED ON WEIGHT CHANGE USING ALL DATA (power 1/%0.1f) --------------------------\n" % power
        powerRegArray = np.array([])
        powerRegArray = pow(sTimeArray,(1/power))
        slope, intercept, r_value, p_value, std_err = stats.linregress(powerRegArray, sy)
        weightRegression.plot(powerRegArray, sy, 'x')
        xls = (np.amin(powerRegArray), np.amax(powerRegArray))
        yls = (np.amin(sy), np.amax(sy))
        x1 = 0.0
        x2 = 0.0
        weightRegression.axis([xls[0], xls[1], yls[0], yls[1]])
        weightRegression.set_xlabel('Time (m^1/4)', fontsize=8)
        weightRegression.set_ylabel('Weight Change (g)', fontsize=8)
        weightRegression.ticklabel_format(style='plain', axis='both, fontsize=8')

        if slope == 0.0:        # the line is parallel to the x axis
            x1 = xls[0]
            x2 = xls[1]
        elif slope > 0.0:
            x1 = max(xls[0], (yls[0] - intercept) / slope)
            x2 = min(xls[1], (yls[1] - intercept) / slope)
        else:
            x1 = max(xls[0], (yls[1] - intercept) / slope)
            x2 = min(xls[1], (yls[0] - intercept) / slope)
        xcoords = np.array([x1, x2])
        ycoords = slope * xcoords + intercept
        weightRegression.plot(xcoords, ycoords, color='r', linewidth=4, label='y = ' + str(slope) + '*x ' +
                                                                              '+-'[intercept < 0.0] + ' ' + str(
            abs(intercept)))
        weightRegression.grid(True)
        results += age_output(power, slope, intercept, r_value, p_value, std_err, ffilename, prefireWeightAverage,
                              prefireWeightStdDev, postfireWeightAverage,
                              postfireWeightStdDev, weightChange, percentWeightChange, runDate, sampleName)
        results += " ------ RESULTS BASED ON WEIGHT CHANGE USING ALL DATA (power %0.3f) ---------------------\n" % power
        results += age_output_nlin(nlinpower, power_stderr, alpha, alpha_stderr, prefireWeightAverage,
                                   prefireWeightStdDev, postfireWeightAverage,
                                   postfireWeightStdDev)
        results += " ------ RESULTS BASED ON WEIGHT CHANGE FILTERED BY MEAN RH +/- 1% -----------------------\n"
        fslope, fintercept, fr_value, fp_value, fstd_err = stats.linregress(fx, fy)
        filteredRHRegressionPlot.plot(fx, fy, 'x')

        xls = (np.amin(fx), np.amax(fx))
        yls = (np.amin(fy), np.amax(fy))
        x1 = 0.0
        x2 = 0.0
        filteredRHRegressionPlot.axis([xls[0], xls[1], yls[0], yls[1]])
        filteredRHRegressionPlot.set_xlabel('Time (m^1/4)', fontsize=8)
        filteredRHRegressionPlot.set_ylabel('Weight Change (g)', fontsize=8)
        filteredRHRegressionPlot.ticklabel_format(style='plain', axis='both, fontsize=8')

        if slope == 0.0:        # the line is parallel to the x axis
            x1 = xls[0]
            x2 = xls[1]
        elif slope > 0.0:
            x1 = max(xls[0], (yls[0] - fintercept) / fslope)
            x2 = min(xls[1], (yls[1] - fintercept) / fslope)
        else:
            x1 = max(xls[0], (yls[1] - fintercept) / fslope)
            x2 = min(xls[1], (yls[0] - fintercept) / fslope)
        xcoords = np.array([x1, x2])
        ycoords = fslope * xcoords + fintercept
        filteredRHRegressionPlot.plot(xcoords, ycoords, color='r', linewidth=4, label='y = ' + str(fslope) + '*x ' +
                                                                                      '+-'[
                                                                                      fintercept < 0.0] + ' ' + str(
            abs(fintercept)))
        filteredRHRegressionPlot.grid(True)
        results += age_output(power, fslope, fintercept, fr_value, fp_value, fstd_err, ffilename, prefireWeightAverage,
                              prefireWeightStdDev, postfireWeightAverage,
                              postfireWeightStdDev, weightChange, percentWeightChange, runDate, sampleName)

        results += " ------ RESULTS BASED ON WEIGHT CHANGE ON EVENLY SPACED POINTS IN TIME^1/4 --------------\n"

        eslope, eintercept, er_value, ep_value, estd_err = stats.linregress(ex, ey)
        evenPlotRegression.plot(ex, ey, 'x')
        xls = (np.amin(ex), np.amax(ex))
        yls = (np.amin(ey), np.amax(ey))
        x1 = 0.0
        x2 = 0.0
        evenPlotRegression.axis([xls[0], xls[1], yls[0], yls[1]])
        evenPlotRegression.set_xlabel('Time (m^1/4)', fontsize=8)
        evenPlotRegression.set_ylabel('Weight Change (g)', fontsize=8)
        evenPlotRegression.ticklabel_format(style='plain', axis='both, fontsize=8')

        if slope == 0.0:        # the line is parallel to the x axis
            x1 = xls[0]
            x2 = xls[1]
        elif slope > 0.0:
            x1 = max(xls[0], (yls[0] - eintercept) / eslope)
            x2 = min(xls[1], (yls[1] - eintercept) / eslope)
        else:
            x1 = max(xls[0], (yls[1] - eintercept) / eslope)
            x2 = min(xls[1], (yls[0] - eintercept) / eslope)
        xcoords = np.array([x1, x2])
        ycoords = eslope * xcoords + eintercept
        evenPlotRegression.plot(xcoords, ycoords, color='r', linewidth=4, label='y = ' + str(eslope) + '*x ' +
                                                                                '+-'[eintercept < 0.0] + ' ' + str(
            abs(eintercept)))
        evenPlotRegression.grid(True)

        results += age_output(power, eslope, eintercept, er_value, ep_value, estd_err, ffilename, prefireWeightAverage,
                              prefireWeightStdDev, postfireWeightAverage,
                              postfireWeightStdDev, weightChange, percentWeightChange, runDate, sampleName)

        if CALCULATE_AGE_FROM_PERCENTS is True:
            results += " ------ RESULTS BASED ON PERCENT WEIGHT CHANGE ---------------------------------------\n"
            pslope, pintercept, pr_value, pp_value, pstd_err = stats.linregress(ppx, ppy)
            pdateYears, pdateCalendar, pADBC, pdateErrorYears, pslopeFourthPower = age_calculateFromPercent(power,
                                                                                                            pslope,
                                                                                                            pstd_err,
                                                                                                            percentWeightChange)
            results += "Percent Weight Change (105 - 500) (%):      " + str(percentWeightChange) + "\n"
            results += "Percent Weight Change RHX Slope (%/m^1/4):  " + str(pslope) + "\n"
            results += "Percent Weight Change RHX Slope (%/m):      " + str(pslopeFourthPower) + "\n"
            results += "Percent Weight Change RHX Intercept (g):    " + str(pintercept) + "\n"
            results += "Percent Weight Change RHX Slope r-value:    " + str(pr_value) + "\n"
            results += "Percent Weight Change RHX Slope std_err:    " + str(pstd_err) + "\n"
            results += "Percent Weight Change Age (Years):          " + str(pdateYears) + " (BP) +/- "\
                       + str(pdateErrorYears) + " " + pADBC + "\n"
            results += "Percent Weight Change Date:                 " + str(abs(pdateCalendar)) + " "\
                       + str(pADBC) + " +/-" + str(pdateErrorYears) + "\n"
            results += "------------------------------------------------------------------------------------------" + "\n"

        plt.show()

        ## if I can figure out how to convert any given temperature to another...
        #ELT = askForELT(sampleNumber,meanTemp)

        #print "ypercent: ", ypercent
        #alphaPercentWeight = nlinRegression(timeArray,ypercent, minval,maxval)
        #minutesPercentWeight = dateCalcFromNlinRegression(alphaPercentWeight,weightChange)
        #dateYears,dateCalendar,ADBC=minToYears(minutesPercentWeight)

        #weightPercentPlotWithLinearTime.plot(timeArray,alphaPercentWeight*np.power(timeArray,0.25)) # 2*sin(x)/x and 3*sin( x)/x

        #results += "Nonlinear Regression Results\n"
        #results += "----------------------------\n"
        #results += "Weight Percent Alpha Value: "+ str(alphaPercentWeight) + " (y = alpha * t^0.25) \n"
        #results += "Weight Percent Minutes: "+ str(minutesPercentWeight) + " (t = (y/alpha)^4)\n"
        #results += "Weight Percent Date in Years BP: "+ str(dateYears) + " (minutes/hours/days = years)\n"
        #results += "Weight Percent Calendar Date: "+ str(dateCalendar) + " "+ ADBC+ "\n"

        #alphaAbsWeight = nlinRegression(timeArray,weightChangeArray, minval,maxval)
        #minutesAbsWeight = dateCalcFromNlinRegression(alphaAbsWeight,percentWeightChange)
        #dateYears,dateCalendar,ADBC=minToYears(minutesAbsWeight)
        #weightPlotWithLinearTime.plot(timeArray,alphaAbsWeight*np.power(timeArray,0.25))

        #results += "----------------------------\n"
        #results += "Abs Weight Alpha Value: "+ str(alphaAbsWeight)+ "\n"
        #results += "Abs Weight  Minutes: "+ str(minutesAbsWeight)+ "\n"
        #results += "Abs Weight  Date in Years BP: "+ str(dateYears)+ "\n"
        #results += "Abs Weight  Calendar Date: "+ str(dateCalendar)+ " "+ ADBC+ "\n"

        print results

        easygui.textbox(msg="Results", title='Results ', text=results)
        test = ""
        if ExcelOrRHX == 1:
            msg = "What would you like to do?"
            choices = ["Try Again", "Quit"]
            reply = easygui.buttonbox(msg, image=None, choices=choices)
        else:
            msg = "What would you like to do?"
            choices = ["Try Again", "Next Sample", "Quit"]
            reply = easygui.buttonbox(msg, image=None, choices=choices)

        if reply == "Next Sample":
            plt.clf()
            sampleNumber += 1
        elif reply == "Try Again":
            plt.clf()
        else:
            if ExcelOrRHX == 2:
                closeDatabase()
            sys.exit()






